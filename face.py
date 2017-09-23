#!/usr/bin/env python2.7
"""
This module contains the graphical user interface and main function.
----------
The GUI class displays a QMainWindow object that interacts with the other
application modules and allows the user to enter and manipulate nutrition and
fitness data. The GUI's primary elements are separated into four pages that are
navigable via Main Menu buttons at the top of the window:

    Home: includes a Login Menu to create, delete, and log in users, a
        Settings button to edit user settings, and a Main Message label that
        displays useful information at the bottom of the page
    Profile: includes the Health Diary and Nutrient Guide Managers, each of
        which allows the user to modify the applicable profile data, and a
        Data Plotter tool that lets the user view and save data plots for
        profile and record data
    Builder: includes the Build Manager tool that lets the user create and
        modify templates and record, the Build Info Viewer that shows current
        nutrition or fitness information for the build, and the Inventory
        Manager tool that lets the user view, edit, and delete reference,
        template, and record inventory items
    Guide: includes an overview of the application tools and directions for
        their use, as well as a glossary of application-specific terms

Each application tool serves a specific purpose and is accompanied by a button
menu with which the user can view, analyze, and modify different types of
application data. These tools are:

    Build Info Viewer: used to view analytical information for the build
        depending on whether it is a nutrition or fitness routine
    Build Manager: used to build nutrition and fitness routines, save build
        elements as templates, and save Diet and Program records--which
        together constitute the bulk of all relevant user data saved to the
        hard drive
    Data Plotter: used to plot progress results from data sources such as the
        Health Diary, Nutrient Guide, Diet records, and Program records
    Health Diary Manager: used to view, add, and delete entries, delete health
        metrics from all entries, and save the Health Diary as a spreadsheet
    Inventory Manager: used to view, add, edit, and delete reference, template,
        and record items, as well as upload and save items as data capsules
    Nutrient Guide Manager: used to view, add, and delete nutrient targets

Button menus have a typical UX-cognizant design. Each button has a status tip
that appears in the status bar when hovered over by the cursor. Menu buttons
are dynamically enabled and disabled depending on whether or not the user can
perform an action. The avoids having to continually warn the user when a menu
action cannot be executed.

This module includes the main function which launches the application. It
checks that the Reference Source directory exists and, if so, creates a lock
file that prevents multiple instances of the application from being open at the
same time. The lock file is closed once the application exits.
----------
Fonts: all QFonts used by this module.
    FONT_TITLE: QFont object for the title
    FONT_TAGA: QFont object for page headers
    FONT_TAGB: QFont object for page subheaders
    FONT_TAGC: QFont object for field descriptors and button text
    FONT_TAGD: QFont object for field descriptors and button text (alternate)
    FONT_TAGE: QFont object for Guide Box text
    FONT_TAGF: QFont object for field descriptors and button text (alternate)

Constants: all constants used by this module.
    INVENTORY_CIDS: dict of inventory IDs mapped to corresponding build element
        class IDs
    INVENTORY_FAVE_IDS: dict of inventory IDs mapped to the corresponding
        favorites ID in settings
    INVENTORY_NAMES: dict of inventory IDs mapped to corresponding inventory
        names

OnTrack GUI: the GUI class and this application's main function.
    MainWindow: QMainWindow subclass for the OnTrack graphical user interface
    main: function to run this application
"""

import datetime
import json
import os
import sys
import ujson

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from PyQt4 import QtCore
from PyQt4 import QtGui

import body
import brain
import dna
import ears
import eyes
import organs


# -----------------------------------------------------------------------------
# Fonts -----------------------------------------------------------------------

FONT_TITLE = organs.arial_huge
FONT_TAGA = organs.arial_large
FONT_TAGB = organs.arial_medium
FONT_TAGC = organs.tahoma_large
FONT_TAGD = organs.arial_small
FONT_TAGE = organs.tahoma_medium
FONT_TAGF = organs.tahoma_small


# -----------------------------------------------------------------------------
# Constants -------------------------------------------------------------------

INVENTORY_CIDS = {
    0: "I", 1: "A", 2: "R", 3: "M", 4: "D", 5: "W", 6: "C", 7: "P", 8: "D",
    9: "P"}
INVENTORY_FAVE_IDS = {
    0: "F", 1: "E", 2: "TR", 3: "TM", 4: "TD", 5: "TW", 6: "TC", 7: "TP",
    8: "RD", 9: "RP"}
INVENTORY_NAMES = {
    0: "FOOD REFERENCE INVENTORY", 1: "EXERCISE REFERENCE INVENTORY",
    2: "RECIPE TEMPLATE INVENTORY", 3: "MEAL TEMPLATE INVENTORY",
    4: "DIET TEMPLATE INVENTORY", 5: "WORKOUT TEMPLATE INVENTORY",
    6: "CYCLE TEMPLATE INVENTORY", 7: "PROGRAM TEMPLATE INVENTORY",
    8: "DIET RECORD INVENTORY", 9: "PROGRAM RECORD INVENTORY"}


# -----------------------------------------------------------------------------
# OnTrack GUI -----------------------------------------------------------------


class MainWindow(QtGui.QMainWindow):
    """
    Class to provide a graphical user interface for the application.

    MainWindow inherits all attributes and methods from its superclass,
    QMainWindow. Its inherited closeEvent method is overridden to prompt the
    user with a confirmation dialog, if designated by user settings, before
    closing the application. Attributes are implemented to store the AppState
    and User objects. The former checks and modifies application state, the
    latter checks and modifies the state of the currently logged in user.
    Additional attributes store the application window's height and width. A
    setup_ui method is implemented to set up the UI at app start and when a
    different user logs in. Multiple attributes are created by the setup_ui
    method, and multiple slot methods are implemented to handle the signals
    from fields, buttons, and other objects. A center method is implemented to
    center the window based on screen resolution size.

    The UI implements a series of interactive button menus and display objects
    throughout the application. A Main Menu at the top of the window presents
    four buttons which the user can use to navigate between four sections or
    'pages': Home, Profile, Builder, and Guide. Home tools let the user manage
    user accounts and change settings. Profile tools let the user view and edit
    Health Diary and Nutrient Guide data, as well as visualize nutrition and
    fitness routine progress on data plots. Builder tools let the user build
    and analyze routines with the Build Manager, as well as view and edit
    reference, template, and record inventories with the Inventory Manager.
    Guide provides a user guide and glossary. The UI's main components are:

    --All Pages--
        Main Menu: navigate to one of four application pages
        Main Message: text displayed at the top of the window that points to
            the Guide Page button or shows the currently logged in username
        Page Viewer: display UI components of the selected application page
    --Home Page--
        Login Menu: create, delete, and log in user accounts
        Settings Button: access and modify the application and user settings
        Home Message: useful text applicable to the logged in user
    --Profile Page--
        Health Diary Manager
            Health Diary Menu: create and delete entries, delete a health
                metric, or save the Health Diary as a spreadsheet
            Health Diary Viewer: view, sort, and select entries
        Nutrient Guide Manager
            Nutrient Guide Menu: create and delete nutrient targets or load
                targets for a selected effective date into the viewer
            Nutrient Guide Viewer: view nutrient targets for an effective date
        Data Plotter Menu: select which data plot dialog to load
    --Builder Page--
        Build Manager
            Build Menu: create, modify, save, and delete routine builds
            Build Viewer: view the build structure and select which build
                element to modify using Build Menu options
            Copied Viewer: view which build element is currently copied
        Build Info Viewer: view analytical data relevant to the build
        Inventory Manager
            Inventory Field: select the inventory type to load
            Search Menu: enter search terms used to filter inventory items,
                displaying only those which contain all terms
            Inventory Viewer: view inventory items and select which item to
                modify using Inventory Menu options
            Inventory Menu: create, modify, and delete reference, template, and
                record inventory items, upload or save items as data capsules,
                and switch between all items and favorite items in the viewer
            Selected Viewer: view which inventory item is currently selected
    --Guide Page--
        Overview Viewer: view an overview of the application and thorough
            instructions for using it
        Glossary Viewer: view terms used throughout the application and their
            definitions

    Methods of the AppState and User objects stored in the appstate and user
    attributes are used to transform user inputs and actions into data sets
    that can be saved to the applicable files in the user's folder. Undo/redo
    action events have not been implemented. Once the application window is
    closed, all unsaved data that has been entered is lost.
    """

    def __init__(self):
        """
        Initialize a MainWindow object and its superclass.

        Assigns an AppState object to the appstate attribute and None to the
        user attribute. Calls the setup_ui method to configure UI components.
        """
        QtGui.QMainWindow.__init__(self)
        self.appstate = brain.AppState(self)
        self.user = None
        self.setup_ui()

    def setup_ui(self):
        """
        Set up the user interface.

        Constructs and modifies every UI object and connects all applicable
        signals to slot methods.
        """
        # ------------------- WINDOW PROPERTIES ------------------- #
        # Set to delete on close and set window properties.
        self.setAttribute(QtCore.Qt.WA_DeleteOnClose)
        wid, hgt = 1400, 900
        self.setFixedSize(wid, hgt)
        self.setWindowTitle("OnTrack -- Nutrition and Fitness Tracker")
        self.setWindowIcon(organs.EasyIcon(dna.ICONS["ontrack"]))
        status_bar = QtGui.QStatusBar()
        self.setStatusBar(status_bar)
        self.center()
        # Set a central widget for Main Menu and Page Viewer.
        central = QtGui.QWidget(self)
        self.setCentralWidget(central)
        # -------------------MAIN MENU
        # Attributes: main_bns, main_mapper
        main_menu = organs.EasyFrame(central, [0, 0, wid, 52], main=True)
        self.main_bns = []
        self.main_mapper = QtCore.QSignalMapper()
        mainbn_info = [
            ("Home", "home", "Go to the Home page"),
            ("Profile", "user_black", "Go to the Profile page"),
            ("Builder", "mountain", "Go to the Builder page"),
            ("Guide", "learn", "Go to the Guide page")]
        mainbn_xpos = 10
        for idx in range(len(mainbn_info)):
            info = mainbn_info[idx]
            button = organs.EasyToolButton(
                main_menu, [mainbn_xpos, 6, 120, 40], text=info[0],
                icon_name=info[1], button_type="main", status_tip=info[2])
            self.main_bns.append(button)
            self.main_mapper.setMapping(button, idx)
            button.clicked.connect(self.main_mapper.map)
            mainbn_xpos += 130
        # Check Home button by default, disable Profile/Builder until login.
        self.main_bns[0].setChecked(True)
        self.main_bns[1].setEnabled(False)
        self.main_bns[2].setEnabled(False)
        self.main_mapper.mapped[int].connect(self.set_page)
        # -------------------MAIN MESSAGE
        # Attributes: main_message
        self.main_message = organs.EasyLabel(
            main_menu, [640, 6, wid - 660, 40], font=FONT_TAGB,
            text="<<< Click 'Guide' to learn how to use OnTrack")
        # -------------------PAGE VIEWER
        # Attributes: pageviewer
        self.pageviewer = QtGui.QStackedWidget(central)
        self.pageviewer.setGeometry(0, 60, wid, hgt - 100)
        pages = []
        for _ in range(4):
            page = QtGui.QWidget()
            pages.append(page)
            self.pageviewer.addWidget(page)
        # Go to Home Page by default.
        self.pageviewer.setCurrentIndex(0)

        # ----------------------- HOME PAGE ----------------------- #
        title_frame = organs.EasyFrame(
            pages[0], [(wid / 2) - (420 / 2), 20, 420, 102],
            style_sheet=dna.STYLE_TITLETAG)
        title_tag = organs.EasyLabel(
            title_frame, [10, 0, 400, 70], text="OnTrack", font=FONT_TITLE)
        subtitle_tag = organs.EasyLabel(
            title_frame, [0, 72, 420, 30],
            text="NUTRITION AND FITNESS TRACKER", font=FONT_TAGB)
        # -------------------LOGIN MENU
        # Attributes: active_user_field, new_user_field, login_bns
        login_menu = organs.EasyFrame(
            pages[0], [(wid / 2) - (520 / 2), 290, 520, 176], login=True)
        active_user_tag = organs.EasyLabel(
            login_menu, [20, 10, 300, 22], text="Active Users", font=FONT_TAGC)
        new_user_tag = organs.EasyLabel(
            login_menu, [20, 90, 300, 22], text="New User", font=FONT_TAGC)
        self.active_user_field = organs.EasyComboBox(
            login_menu, [20, 40, 300, 22], items=self.appstate.users(),
            first_item="-" * 14 + "SELECT A USERNAME" + "-" * 14)
        self.new_user_field = organs.EasyLineEdit(
            login_menu, [20, 120, 300, 22], max_length=30,
            placeholder_text="Enter a username up to 30 characters long")
        self.login_bns = []
        loginbn_info = [
            ([340, 20], "Login User", "fingerprint",
             "Load the selected Active User"),
            ([426, 20], "Delete User", "delete",
             "Delete the selected Active User"),
            ([340, 100], "Create User", "user_add",
             "Create an account for the New User")]
        for idx in range(len(loginbn_info)):
            info = loginbn_info[idx]
            button = organs.EasyToolButton(
                login_menu, info[0] + [80, 60], text=info[1], font=FONT_TAGF,
                icon_name=info[2], icon_size=[32, 32],
                toolbutton_style=QtCore.Qt.ToolButtonTextUnderIcon,
                button_type="action", status_tip=info[3], disable=True)
            self.login_bns.append(button)
        self.active_user_field.currentIndexChanged.connect(self.toggle_login)
        self.new_user_field.textChanged.connect(self.toggle_login)
        self.new_user_field.returnPressed.connect(self.create_user)
        self.login_bns[0].clicked.connect(self.login_user)
        self.login_bns[1].clicked.connect(self.delete_user)
        self.login_bns[2].clicked.connect(self.create_user)
        # -------------------SETTINGS
        # Attributes: settings_bn
        self.settings_bn = organs.EasyToolButton(
            pages[0], [(wid / 2) - (80 / 2), 480, 80, 80],
            text="Settings", icon_name="settings", icon_size=[50, 50],
            toolbutton_style=QtCore.Qt.ToolButtonTextUnderIcon,
            button_type="action", status_tip="Change application settings")
        self.settings_bn.clicked.connect(self.edit_settings)
        # -------------------HOME MESSAGE
        # Attributes: home_message
        home_msg = "".join([
            "If this is your first time using OnTrack, you must create a ",
            "new user account before logging in."])
        self.home_message = organs.EasyLabel(
            pages[0], [50, hgt - 150, wid - 100, 30], text=home_msg,
            font=FONT_TAGC, style_sheet=dna.STYLE_HOMEMESSAGE)

        # ---------------------- PROFILE PAGE --------------------- #
        # Profile has a background frame that looks like a file folder.
        pwid, phgt = wid - 20, hgt - 150
        profile_bg = organs.EasyFrame(
            pages[1], [10, 46, pwid, phgt], profile=True)
        profile_tag = organs.EasyLabel(
            pages[1], [10, 0, 250, 47], text="PROFILE",
            font=FONT_TAGA, style_sheet=dna.STYLE_PROFILETAG)
        health_diary_tag = organs.EasyLabel(
            profile_bg, [10, 10, pwid - 700, 28], text="HEALTH DIARY",
            font=FONT_TAGB, style_sheet=dna.STYLE_STICKER)
        nutrient_guide_tag = organs.EasyLabel(
            profile_bg, [pwid - 680, 10, 410, 28], text="NUTRIENT GUIDE",
            font=FONT_TAGB, style_sheet=dna.STYLE_STICKER)
        data_plotter_tag = organs.EasyLabel(
            profile_bg, [pwid - 260, 10, 250, 28], text="DATA PLOTTER",
            font=FONT_TAGB, style_sheet=dna.STYLE_STICKER)
        # -------------------HEALTH DIARY MANAGER
        # Attributes: add_entry_bn, delete_entry_bn, save_diary_bn,
        #             metric_field, delete_metric_bn, diaryviewer
        diary_menu = organs.EasyFrame(profile_bg, [15, 44, pwid - 710, 28])
        # Disable all buttons by default until login.
        self.add_entry_bn = organs.EasyToolButton(
            diary_menu, [0, 0, 130, 28], text="Add Entry", font=FONT_TAGD,
            icon_name=dna.ICONS["diary"], icon_size=[30, 30],
            toolbutton_style=QtCore.Qt.ToolButtonTextBesideIcon, disable=True,
            button_type="action", status_tip="Add a new Health Diary entry")
        self.delete_entry_bn = organs.EasyToolButton(
            diary_menu, [132, 0, 130, 28], text="Delete Entry",
            font=FONT_TAGD, icon_name="delete", icon_size=[30, 30],
            toolbutton_style=QtCore.Qt.ToolButtonTextBesideIcon,
            button_type="action", disable=True,
            status_tip="Delete the selected Health Diary entries")
        self.save_diary_bn = organs.EasyToolButton(
            diary_menu, [264, 0, 130, 28], text="Save Diary",
            font=FONT_TAGD, icon_name="file_download", icon_size=[30, 30],
            toolbutton_style=QtCore.Qt.ToolButtonTextBesideIcon,
            button_type="action", disable=True,
            status_tip="Save your Health Diary as a spreadsheet")
        self.metric_field = organs.EasyComboBox(
            diary_menu, [pwid - 970, 0, 220, 28],
            style_sheet=dna.STYLE_COMBOBOX, tool_tip="Health Metric",
            status_tip="Select a Health Metric to delete")
        self.delete_metric_bn = organs.EasyToolButton(
            diary_menu, [pwid - 750, 0, 40, 28], icon_name="delete",
            icon_size=[30, 30], toolbutton_style=QtCore.Qt.ToolButtonIconOnly,
            button_type="action", disable=True, tool_tip="Delete Metric",
            status_tip="Delete the selected health metric from your diary")
        self.diaryviewer = organs.EasyTableWidget(
            profile_bg, [10, 82, pwid - 700, phgt - 102], row_height=28,
            row_header_hidden=True, edit_off=True, select_rows=True,
            sorting=True)
        self.add_entry_bn.clicked.connect(self.add_entry)
        self.delete_entry_bn.clicked.connect(self.delete_entry)
        self.save_diary_bn.clicked.connect(self.save_diary)
        self.metric_field.currentIndexChanged.connect(self.toggle_diary)
        self.delete_metric_bn.clicked.connect(self.delete_metric)
        self.diaryviewer.itemSelectionChanged.connect(self.toggle_diary)
        # -------------------NUTRIENT GUIDE MANAGER
        # Attributes: add_targets_bn, effective_field, view_targets_bn,
        #             delete_targets_bn, effective_tag, guideviewer
        guide_menu = organs.EasyFrame(profile_bg, [pwid - 675, 44, 400, 28])
        # Disable all buttons by default until login.
        self.add_targets_bn = organs.EasyToolButton(
            guide_menu, [0, 0, 130, 28], text="Add Targets",
            font=FONT_TAGD, icon_name=dna.ICONS["guide"],
            icon_size=[30, 30], button_type="action", disable=True,
            toolbutton_style=QtCore.Qt.ToolButtonTextBesideIcon,
            status_tip="Add new nutrient targets to your Nutrient Guide")
        self.effective_field = organs.EasyComboBox(
            guide_menu, [190, 0, 130, 28], style_sheet=dna.STYLE_COMBOBOX,
            tool_tip="Effective Date",
            status_tip="Select an effective date for which to load targets")
        self.view_targets_bn = organs.EasyToolButton(
            guide_menu, [320, 0, 40, 28], icon_name="refresh", disable=True,
            icon_size=[30, 30], button_type="action", tool_tip="View Targets",
            toolbutton_style=QtCore.Qt.ToolButtonIconOnly,
            status_tip="View targets for the selected effective date")
        self.delete_targets_bn = organs.EasyToolButton(
            guide_menu, [360, 0, 40, 28], icon_name="delete", disable=True,
            button_type="action", tool_tip="Delete Targets",
            toolbutton_style=QtCore.Qt.ToolButtonIconOnly, icon_size=[30, 30],
            status_tip="Delete targets for the selected effective date")
        self.effective_tag = organs.EasyLabel(
            profile_bg, [pwid - 680, 80, 410, 28], text="NO TARGETS SET",
            font=FONT_TAGD)
        self.guideviewer = organs.EasyTableWidget(
            profile_bg, [pwid - 680, 102, 410, phgt - 122], row_height=28,
            edit_off=True, resize_off=True, stretch_last_section=True)
        self.add_targets_bn.clicked.connect(self.add_targets)
        self.effective_field.currentIndexChanged.connect(self.toggle_guide)
        self.view_targets_bn.clicked.connect(self.view_targets)
        self.delete_targets_bn.clicked.connect(self.delete_targets)
        # -------------------DATA PLOTTER
        # Attributes: plot_bns, plot_mapper
        plot_menu = organs.EasyFrame(profile_bg, [pwid - 255, 44, 240, 400])
        self.plot_bns = []
        self.plot_mapper = QtCore.QSignalMapper()
        plotbn_info = [
            (10, "Health Metrics", "diary", "Plot health metric measurements"),
            (42, "Nutrient Targets", "guide", "Plot nutrient targets"),
            (85, "Macronutrient Proportions", "nutrition",
             "Plot protein, fat, and carbohydrate proportions in your diet"),
            (117, "Meal Times", "nutrition", "Plot meal times"),
            (149, "Nutrient Values", "nutrition", "Plot nutrient consumption"),
            (194, "Performance Per Exercise", "fitness",
             "Plot performance results per exercise"),
            (226, "Performance Per Muscle", "fitness",
             "Plot performance results per muscle"),
            (258, "Performance Per Tag", "fitness",
             "Plot performance results per exercise tag"),
            (290, "Workout Periods", "fitness", "Plot workout periods")]
        for idx in range(len(plotbn_info)):
            info = plotbn_info[idx]
            button = organs.EasyToolButton(
                plot_menu, [0, info[0], 240, 30], text=info[1], font=FONT_TAGD,
                icon_name=dna.ICONS[info[2]], icon_size=[32, 32],
                toolbutton_style=QtCore.Qt.ToolButtonTextBesideIcon,
                button_type="action", status_tip=info[3], disable=True)
            self.plot_bns.append(button)
            self.plot_mapper.setMapping(button, idx)
            button.clicked.connect(self.plot_mapper.map)
        self.plot_mapper.mapped[int].connect(self.plot_action)

        # --------------------- BUILDER PAGE  --------------------- #
        bwid, bhgt = 650, hgt - 312
        iwid, ihgt = wid - bwid - 30, hgt - 384
        build_manager_tag = organs.EasyLabel(
            pages[2], [10, 0, bwid, 30], text="BUILD MANAGER",
            font=FONT_TAGB, style_sheet=dna.STYLE_BUILDMANAGERTAG)
        inventory_manager_tag = organs.EasyLabel(
            pages[2], [bwid + 20, 0, iwid, 30], text="INVENTORY MANAGER",
            font=FONT_TAGB, style_sheet=dna.STYLE_INVENTORYMANAGERTAG)
        # -------------------BUILD MANAGER
        # Attributes: copied_element, expanding_message, build_bns,
        #             build_mapper, buildviewer, copiedviewer
        build_menu = organs.EasyFrame(pages[2], [10, 70, 40, bhgt - 30])
        self.copied_element = None
        self.expanding_message = None
        self.build_bns = []
        self.build_mapper = QtCore.QSignalMapper()
        buildbn_info = [
            (0, "add_circleblack", "Create a new nutrition or fitness build",
             "Create New Build"),
            (45, "triangle_up",
             "Move the selected build element one position up",
             "Move Element Up"),
            (60, "triangle_down",
             "Move the selected build element one position down",
             "Move Element Down"),
            (85, "add", "Add a new element to the selected build element",
             "Add New Element"),
            (120, "list_add", "Add the selected inventory item to the " +
             "selected build element", "Add Inventory Item"),
            (155, "edit", "Edit the selected build element's properties",
             "Edit Element"),
            (190, "clear", "Remove the selected build element from the build",
             "Remove Element"),
            (235, "copy", "Copy the selected build element", "Copy Element"),
            (270, "paste",
             "Paste the copied element to the selected build element",
             "Paste Element"),
            (315, "sub_right",
             "Expand all elements under the selected build element",
             "Expand Children"),
            (350, "sub_left",
             "Collapse all elements under the selected build element",
             "Collapse Children"),
            (395, "star_black",
             "Save the selected build element as a template",
             "Save Element Template"),
            (430, "save", "Save your Diet or Program build as a record",
             "Save Build Record"),
            (475, "clear_circleblack",
             "Clear your build from the Build Viewer", "Clear Build Viewer")]
        for idx in range(len(buildbn_info)):
            info = buildbn_info[idx]
            bn_height = 35 if idx not in [1, 2] else 15
            button = organs.EasyToolButton(
                build_menu, [0, info[0], 35, bn_height],
                icon_name=info[1], icon_size=[34, 34], disable=True,
                toolbutton_style=QtCore.Qt.ToolButtonIconOnly,
                button_type="action", status_tip=info[2], tool_tip=info[3])
            self.build_bns.append(button)
            self.build_mapper.setMapping(button, idx)
            button.clicked.connect(self.build_mapper.map)
        self.buildviewer = organs.BuildViewer(
            pages[2], [50, 40, bwid - 40, bhgt])
        copied_tag = organs.EasyLabel(
            pages[2], [50, bhgt + 46, 80, 24], text="COPIED >",
            font=FONT_TAGD, style_sheet=dna.STYLE_COPYTAG)
        self.copiedviewer = organs.EasyLabel(
            pages[2], [132, bhgt + 46, bwid - 122, 24],
            style_sheet=dna.STYLE_STICKER,
            alignment=QtCore.Qt.AlignLeft | QtCore.Qt.AlignVCenter)
        self.build_mapper.mapped[int].connect(self.build_action)
        self.buildviewer.currentItemChanged.connect(self.toggle_build)
        # -------------------BUILD INFO VIEWER
        # Attributes: buildinfoviewer, nutritioninfo, fitnessinfo
        self.buildinfoviewer = QtGui.QStackedWidget(pages[2])
        self.buildinfoviewer.setGeometry(10, hgt - 232, wid - 20, 132)
        self.buildinfoviewer.setFrameShape(QtGui.QFrame.Box)
        infopages = []
        for _ in range(3):
            page = QtGui.QWidget()
            infopages.append(page)
            self.buildinfoviewer.addWidget(page)
        self.buildinfoviewer.setCurrentIndex(0)
        blank_tag = organs.EasyLabel(
            infopages[0], [10, 10, wid - 20, 110], text="BUILD INFO",
            font=FONT_TAGC)
        self.nutritioninfo = organs.EasyTableWidget(
            infopages[1], [0, 0, wid - 22, 130], column_width=82,
            dims=[0, 3], row_labels=["VALUE", "UNIT", "% TARGET"],
            row_height=24, row_label_alignment=QtCore.Qt.AlignCenter,
            edit_off=True, resize_rows_off=True, single_selection=True)
        self.fitnessinfo = organs.EasyTableWidget(
            infopages[2], [0, 0, wid - 22, 130], column_width=110,
            row_labels=["TOTAL SESSIONS", "TOTAL EFFORT", "MAX INTENSITY"],
            dims=[0, 3], row_label_alignment=QtCore.Qt.AlignCenter,
            row_height=24, edit_off=True, resize_rows_off=True,
            single_selection=True)
        # -------------------INVENTORY MANAGER
        # Attributes: view_favorites, inventory_tag, inventory_field,
        #             search_field, search_bn, inventoryviewer, inventory_bns,
        #             inventory_mapper, selectedviewer
        self.view_favorites = False
        self.inventory_tag = organs.EasyLabel(
            pages[2], [bwid + 40, 38, iwid, 30], text="", font=FONT_TAGC,
            alignment=QtCore.Qt.AlignVCenter | QtCore.Qt.AlignLeft)
        inventory_names = [
            "Reference: Food", "Reference: Exercise",
            "Template: Recipe", "Template: Meal", "Template: Diet",
            "Template: Workout", "Template: Cycle", "Template: Program",
            "Record: Diet", "Record: Program"]
        self.inventory_field = organs.EasyComboBox(
            pages[2], [bwid + iwid - 180, 34, 180, 30], items=inventory_names,
            style_sheet=dna.STYLE_COMBOBOX, tool_tip="Select Inventory",
            status_tip="Select which inventory to load")
        place = "".join([
            "Filter the inventory by search terms, delete search terms to ",
            "reset filter and view all items"])
        self.search_field = organs.EasyLineEdit(
            pages[2], [bwid + 20, 76, iwid - 72, 32], placeholder_text=place)
        self.search_bn = organs.EasyToolButton(
            pages[2], [bwid + iwid - 52, 76, 32, 32], icon_name="search",
            icon_size=[30, 30], button_type="action",
            toolbutton_style=QtCore.Qt.ToolButtonIconOnly,
            status_tip="Filter the inventory items by search terms",
            tool_tip="Search Inventory", disable=True)
        self.inventoryviewer = organs.EasyTableWidget(
            pages[2], [bwid + 20, 112, iwid - 40, ihgt], dims=[1, 0],
            column_labels=["Inventory Viewer"], row_height=22,
            row_header_hidden=True, edit_off=True, grid_off=True,
            select_rows=True, single_selection=True, sorting=True)
        inventory_menu = organs.EasyFrame(
            pages[2], [bwid + iwid - 20, 112, 40, ihgt])
        self.inventory_bns = []
        self.inventory_mapper = QtCore.QSignalMapper()
        inventoryviewerbn_info = [
            (0, "go_left", "Send the selected template or record to the " +
             "Build Viewer and start a build", "Send Item To Build"),
            (45, "add", "Create a new Food or Exercise reference item",
             "Create New Reference"),
            (80, "file_upload", "Create new Food or Exercise reference " +
             "item(s) from 'Data Capsule' files on your hard drive",
             "Load Data Capsules"),
            (125, "visible_on", "View the selected inventory item's " +
             "properties--double clicking an item works, too", "View Item"),
            (160, "edit", "Edit the selected inventory item's properties",
             "Edit Item"),
            (195, "file_download", "Save the selected reference item as a " +
             "shareable 'Data Capsule' file", "Save Data Capsule"),
            (230, "delete", "Delete the selected inventory item",
             "Delete Item"),
            (275, "list", "View all inventory items", "View All Items"),
            (310, "heart_black", "View favorite inventory items",
             "View Favorite Items")]
        for idx in range(len(inventoryviewerbn_info)):
            info = inventoryviewerbn_info[idx]
            bn_type = "selection" if idx in [7, 8] else "action"
            button = organs.EasyToolButton(
                inventory_menu, [5, info[0], 35, 35], icon_name=info[1],
                icon_size=[34, 34], button_type=bn_type, status_tip=info[2],
                tool_tip=info[3], disable=True,
                toolbutton_style=QtCore.Qt.ToolButtonIconOnly)
            self.inventory_bns.append(button)
            self.inventory_mapper.setMapping(button, idx)
            button.clicked.connect(self.inventory_mapper.map)
        selected_tag = organs.EasyLabel(
            pages[2], [bwid + 20, bhgt + 46, 110, 24],
            text="SELECTED >", font=FONT_TAGD, style_sheet=dna.STYLE_SELECTTAG)
        self.selectedviewer = organs.EasyLabel(
            pages[2], [bwid + 132, bhgt + 46, iwid - 152, 24],
            style_sheet=dna.STYLE_STICKER,
            alignment=QtCore.Qt.AlignLeft | QtCore.Qt.AlignVCenter)
        self.inventory_field.currentIndexChanged.connect(self.load_inventory)
        self.search_field.textChanged.connect(self.toggle_search)
        self.search_field.returnPressed.connect(self.search_inventory)
        self.search_bn.clicked.connect(self.search_inventory)
        self.inventoryviewer.itemSelectionChanged.connect(
            self.toggle_inventory)
        self.inventoryviewer.itemSelectionChanged.connect(self.toggle_build)
        self.inventoryviewer.itemDoubleClicked.connect(self.view_item)
        self.inventory_mapper.mapped[int].connect(self.inventory_action)

        # ----------------------- GUIDE PAGE ---------------------- #
        ghgt = hgt - 150
        guide_tag = organs.EasyLabel(
            pages[3], [10, 0, 250, 44], text="GUIDE", font=FONT_TAGA,
            style_sheet=dna.STYLE_GUIDETAG)
        guide_box_a = organs.TipBox(
            pages[3], [10, 50, (wid * 0.5) - 15, ghgt], text=dna.GUIDE_TEXTA,
            guide=True, font=FONT_TAGE)
        guide_box_b = organs.TipBox(
            pages[3], [(wid * 0.5) + 5, 50, (wid * 0.5) - 15, ghgt],
            text=dna.GUIDE_TEXTB, guide=True, font=FONT_TAGE)

        # ---------------------- DEFAULT USER --------------------- #
        QtGui.QApplication.processEvents()
        default_user = self.appstate.default_user()
        if default_user is not None:
            self.active_user_field.blockSignals(True)
            self.active_user_field.setCurrentIndex(
                self.active_user_field.findText(default_user))
            self.active_user_field.blockSignals(False)
            self.login_user()

    # ------------------------- MAIN MENU ------------------------- #
    def set_page(self, page_num):
        """
        Set the application page.

        :param page_num: Page number to switch to: 0-Home, 1-Profile,
            2-Builder, 3-Guide
        :type page_num: int
        """
        self.pageviewer.setCurrentIndex(page_num)

    # ------------------------- LOGIN MENU ------------------------ #
    def login_user(self):
        """
        Set up the UI for user selected in the Active User Field.

        Updates the UI for the username selected in the Active User Field.
        If the user attribute is not None, prompts the current user to confirm
        logging out. Instantiates a new User object for the selected user and
        assigned it to the user attribute. Updates the UI accordingly.
        """
        selected_user = str(self.active_user_field.currentText())
        # Check for logged in user.
        if self.user is not None:
            current_user = self.user.username()
            choice = QtGui.QMessageBox.warning(
                self, "Confirm Log Out", "Are you sure you want to log out " +
                "and lose any unsaved data?",
                QtGui.QMessageBox.Yes | QtGui.QMessageBox.No)
            if choice == QtGui.QMessageBox.No:
                return
        # Disable Profile/Builder buttons and clear Main/Home Messages.
        self.main_bns[1].setEnabled(False)
        self.main_bns[2].setEnabled(False)
        self.main_message.setText("Logged In: ---")
        self.home_message.setText("Loading...")
        # Display loading message and process events.
        loading_message = organs.MessageDialog(
            self, "Loading User", "LOGGING YOU IN...")
        loading_message.show()
        QtGui.QApplication.processEvents()
        # Set user attribute.
        self.user = brain.User(self, selected_user)
        self.toggle_login()
        # Rebuild and toggle Profile tools.
        self.update_diary()
        self.update_guide()
        self.toggle_plot()
        # Rebuild and toggle Build Manager components.
        self.buildviewer.clear_buildparent()
        self.copied_element = None
        self.copiedviewer.clear()
        self.update_buildinfo(rebuild=True)
        self.buildinfoviewer.setCurrentIndex(0)
        # Rebuild and toggle Inventory Manager components. Toggling Inventory
        # Menu also toggles Build Menu.
        self.view_favorites = False
        self.search_field.clear()
        self.load_inventory(0)
        self.inventory_field.setCurrentIndex(0)
        # Enable Profile/Builder buttons and set Main/Home Messages.
        self.main_bns[1].setEnabled(True)
        self.main_bns[2].setEnabled(True)
        self.main_message.setText("Logged In:  " + selected_user)
        home_msg = "".join([
            "Welcome back! Check the Guide Page if you have any questions."])
        self.home_message.setText(home_msg)
        # Close loading message.
        loading_message.finish()

    def delete_user(self):
        """
        Delete the user selected in the Active User Field.

        Prompts the user to confirm deletion of the selected user. If the user
        attribute is not None, its applicable method is called to remove the
        user files and folder from the Users directory. If user is None, a
        temporary User object is created for the selected user and its method
        is called. The UI, appstate attribute, and Active User Field are
        updated accordingly.
        """
        selected_user = str(self.active_user_field.currentText())
        # Check if user wants to delete this username.
        if self.user is not None and selected_user == self.user.username():
            choice = QtGui.QMessageBox.warning(
                self, "Confirm Delete User", "Are you sure you want to " +
                "delete all data for the currently logged in user:\n\n" +
                selected_user, QtGui.QMessageBox.Yes | QtGui.QMessageBox.No)
            if choice == QtGui.QMessageBox.No:
                return
            self.user.delete_user_directory()
            self.user = None
            self.main_bns[1].setEnabled(False)
            self.main_bns[2].setEnabled(False)
            self.main_message.setText("Logged In: ---")
            home_msg = "".join([
                "You are no longer logged in. Create a new user, or log in ",
                "another active user."])
            self.home_message.setText(home_msg)
        else:
            choice = QtGui.QMessageBox.warning(
                self, "Confirm Delete User", "Are you sure you want to " +
                "delete all data for the user:\n\n" + selected_user,
                QtGui.QMessageBox.Yes | QtGui.QMessageBox.No)
            if choice == QtGui.QMessageBox.No:
                return
            temp_user = brain.User(self, selected_user)
            temp_user.delete_user_directory()
        # Update appstate, Active User Field, and default user, if applicable.
        self.appstate.remove_user(selected_user)
        if self.appstate.default_user() == selected_user:
            self.appstate.set_default_user(None)
        self.active_user_field.removeItem(
            self.active_user_field.findText(selected_user))
        self.active_user_field.setCurrentIndex(0)
        self.toggle_login()

    def create_user(self):
        """
        Create user files for the username entered into the New User Field.

        Verifies that the username does not already exist and that it contains
        only valid characters (letters, numbers, underscores, hyphens). A
        temporary User object is instantiated for the username so that user
        files and a folder can be created in the User directory. If the user is
        created without exceptions being raised, the user is notified that the
        new user has been created. Updates appstate attribute and Active User
        Field.
        """
        new_user = str(self.new_user_field.text()).strip()
        # Check for existing username (case-insensitive).
        for username in self.appstate.users():
            if username.lower() == new_user.lower():
                QtGui.QMessageBox.warning(
                    self, "Existing Username", "This user already exists! " +
                    "Be aware that usernames are not case-sensitive.")
                return
        # Check for only alphanumeric, underscore, and hyphen characters.
        legal_chars = "ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz"
        legal_chars += "0123456789-_"
        for char in new_user:
            if char not in legal_chars:
                QtGui.QMessageBox.warning(
                    self, "Invalid Username",
                    "Your new username must be any combination of letters, " +
                    "numbers, hyphens, and underscores (with no spaces)!\n\n" +
                    "Example: Got_Apples--2017")
                return
        # Clear New User Field. If no exceptions raised when creating user,
        # update appstate and Active User Field.
        self.new_user_field.clear()
        temp_user = brain.User(self, new_user, new_user=True)
        if temp_user.was_created():
            self.appstate.add_user(new_user)
            self.active_user_field.addItem(new_user)
            QtGui.QMessageBox.information(
                self, "Created New User",
                "You have successfully created the user:\n\n" + new_user)

    def toggle_login(self):
        """
        Toggle the Login Menu buttons.

        Enables or disables Login Menu buttons based on the user attribute,
        Active User Field, and New User Field values.
        """
        current = "" if self.user is None else self.user.username()
        placeholder = "-" * 14 + "SELECT A USERNAME" + "-" * 14
        active = str(self.active_user_field.currentText())
        new = str(self.new_user_field.text()).strip()
        enable = []
        disable = []
        # Enable Login User if active username is selected and not logged in.
        action = enable if active not in [current, placeholder] else disable
        action.append(self.login_bns[0])
        # Enable Delete User if active username is selected.
        action = enable if active != placeholder else disable
        action.append(self.login_bns[1])
        # Enable Create User if new username is entered.
        action = enable if new != "" else disable
        action.append(self.login_bns[2])
        # Enable and disable buttons.
        for button in enable:
            button.setEnabled(True)
        for button in disable:
            button.setEnabled(False)

    def edit_settings(self):
        """
        Edit application and user settings.

        The app_state attribute's default user is always replaced with the
        dialog's default_user value. If a user is logged in, and thus given
        user settings to edit, the user's settings are updated by passing the
        dialog.settings value to its update_settings method. The Build Info is
        also updated in case of an active build.
        """
        dialog = ears.InputSettings(self, self.appstate, self.user)
        if dialog.exec_():
            self.appstate.set_default_user(dialog.default_user)
            if self.user is not None:
                self.user.update_settings(dialog.settings)
                self.update_buildinfo(rebuild=True)

    # --------------------- HEALTH DIARY MENU --------------------- #
    def update_diary(self):
        """
        Update the Health Diary Viewer.

        Clears the Health Diary Viewer table and repopulates it with the user's
        current Health Diary entries. If no entries exist, the table will be
        empty. The table format is:

            'Date'              First Health Metric     Second Health Metric
            First Entry Date    measurement value       measurement value
            Second Entry Date   measurement value       measurement value

        Each time the table is populated, the rows are automatically sorted by
        the 'Date' column's entry date values.
        """
        self.diaryviewer.blockSignals(True)
        self.diaryviewer.setRowCount(0)
        self.diaryviewer.setColumnCount(0)
        self.metric_field.blockSignals(True)
        self.metric_field.clear()
        self.metric_field.addItem("-----CHOOSE A METRIC-----")
        # If the Health Diary has entries, add them to the table.
        health_diary = self.user.health_diary()
        if not health_diary:
            self.diaryviewer.blockSignals(False)
            self.metric_field.blockSignals(False)
            self.toggle_diary()
            return
        sorted_metrics = organs.unique_keys(health_diary, 1)
        sorted_entry_dates = sorted(health_diary.keys())
        self.diaryviewer.setColumnCount(len(sorted_metrics) + 1)
        self.diaryviewer.setRowCount(len(sorted_entry_dates))
        self.diaryviewer.setHorizontalHeaderLabels(["Date"] + sorted_metrics)
        for entry_date in sorted_entry_dates:
            # Center all cell text and make dates bold (tag D).
            row = sorted_entry_dates.index(entry_date)
            date_item = QtGui.QTableWidgetItem()
            date_item.setText(entry_date)
            date_item.setTextAlignment(QtCore.Qt.AlignCenter)
            date_item.setFont(FONT_TAGD)
            self.diaryviewer.setItem(row, 0, date_item)
            # Only add QTableWidgetItem objects to cells which have values.
            for metric in health_diary[entry_date]:
                col = sorted_metrics.index(metric)
                value_item = QtGui.QTableWidgetItem()
                value_item.setText(str(health_diary[entry_date][metric]))
                value_item.setTextAlignment(QtCore.Qt.AlignCenter)
                self.diaryviewer.setItem(row, col + 1, value_item)
        # Add all metrics to the Health Metric Field.
        self.metric_field.addItems(sorted_metrics)
        self.diaryviewer.blockSignals(False)
        self.metric_field.blockSignals(False)
        self.toggle_diary()

    def add_entry(self):
        """Add an entry to the Health Diary."""
        dialog = ears.InputEntry(self, self.user)
        if dialog.exec_():
            self.user.add_entry(dialog.entry)
            self.update_diary()
            self.toggle_plot()

    def delete_entry(self):
        """Delete all selected entries from the Health Diary."""
        # Find the row numbers of all selected items.
        rows = set([item.row() for item in self.diaryviewer.selectedItems()])
        pronoun = "this entry?" if len(rows) == 1 else "these entries?"
        choice = QtGui.QMessageBox.warning(
            self, "Confirm Delete Entries", "Are you sure you want to " +
            "delete " + pronoun, QtGui.QMessageBox.Yes | QtGui.QMessageBox.No)
        if choice == QtGui.QMessageBox.No:
            return
        for row in rows:
            entry_date = str(self.diaryviewer.item(row, 0).text())
            self.user.remove_entry(entry_date)
        self.update_diary()
        self.toggle_plot()

    def save_diary(self):
        """
        Save the Health Diary as a spreadsheet.

        Parses the Health Diary and creates a spreadsheet that can be saved as
        a .xlsx file. Prompts the user to selected the save directory. The xlsx
        file contains a single worksheet with all entry data. The first row
        includes column headers for the 'Date' followed by each health metric.
        Under the header, the first column lists each entry dated sorted in
        ascending order, and each subsequent column lists measurement values
        for the corresponding health metrics. Blank cells in this grid indicate
        that no measurements were recorded. The table format is:

            'Date'              First Health Metric     Second Health Metric
            First Entry Date    measurement value       measurement value
            Second Entry Date   measurement value       measurement value

        Calls EasyFileDialog with the

            <USERNAME>'s Health Diary_<CURRENT_DATE>.xlsx
        """
        try:
            health_diary = self.user.health_diary()
            sorted_entry_dates = sorted(health_diary.keys())
            sorted_metrics = organs.unique_keys(health_diary, 1)
            sorted_metrics.insert(0, "Date")
            # Build the spreadsheet. Name the tab "Health Diary".
            workbook = Workbook()
            wba = workbook.active
            wba.title = "Health Diary"
            column_widths = [0] * len(sorted_metrics)
            for row in range(1, len(sorted_entry_dates) + 2):
                if row > 1:
                    date = sorted_entry_dates[row - 2]
                for col in range(1, len(sorted_metrics) + 1):
                    metric = sorted_metrics[col - 1]
                    # Check for header.
                    if row == 1:
                        val = metric
                    # Check for date.
                    elif col == 1:
                        val = date
                    # Check for no measurement.
                    elif metric not in health_diary[date]:
                        continue
                    # Use measurement value.
                    else:
                        val = health_diary[date][metric]
                    _ = wba.cell(column=col, row=row, value=val)
                    # Center cell.
                    cell_idx = get_column_letter(col) + str(row)
                    cell = wba[cell_idx]
                    cell.alignment = cell.alignment.copy(
                        horizontal="center", vertical="center")
                    # Check max chars for column width adjustment.
                    if len(str(val)) > column_widths[col - 1]:
                        column_widths[col - 1] = len(str(val))
            # Set column widths.
            for idx, col_wid in enumerate(column_widths):
                wba.column_dimensions[
                    get_column_letter(idx + 1)].width = col_wid + 2
            # Open save file dialog.
            filename = self.user.username() + "'s Health Diary_"
            filename += str(datetime.datetime.now().date())
            dialog = organs.EasyFileDialog(self, "healthdiary", filename)
            if dialog.exec_():
                dest_path = os.path.normpath(str(dialog.selectedFiles()[0]))
                if not dest_path[-5:] == ".xlsx":
                    dest_path += ".xlsx"
                if os.path.isfile(dest_path):
                    choice = QtGui.QMessageBox.warning(
                        self, "Confirm Overwrite File", "This file already " +
                        "exists! Do you want to replace it?",
                        QtGui.QMessageBox.Yes | QtGui.QMessageBox.No)
                    if choice == QtGui.QMessageBox.No:
                        return
                workbook.save(filename=dest_path)
                QtGui.QMessageBox.information(
                    self, "Health Diary Saved",
                    "Your Health Diary has been saved!")
        except IOError:
            QtGui.QMessageBox.warning(
                self, "File Error", "An error occurred while trying to save " +
                "your Health Diary! Make sure you are not trying to save " +
                "over an open file.")

    def delete_metric(self):
        """Delete the selected health metric from the Health Diary."""
        metric = str(self.metric_field.currentText())
        choice = QtGui.QMessageBox.warning(
            self, "Confirm Delete Metric",
            "Are you sure you want to delete the health metric:\n\n" + metric,
            QtGui.QMessageBox.Yes | QtGui.QMessageBox.No)
        if choice == QtGui.QMessageBox.No:
            return
        self.user.remove_health_metric(metric)
        self.update_diary()
        self.toggle_plot()

    def toggle_diary(self):
        """
        Toggle the Health Diary Menu.

        Enables or disables Health Diary Menu buttons based on the user
        attribute, the size of the user's Health Diary (if applicable), if
        entries are selected in the Health Diary Viewer, and the Health Metric
        Field value.
        """
        has_entries = True if self.user and self.user.health_diary() else False
        rows = set([item.row() for item in self.diaryviewer.selectedItems()])
        enable = []
        disable = []
        # Enable Add Entry if user is logged in.
        action = enable if self.user is not None else disable
        action.append(self.add_entry_bn)
        # Enable Delete Entry if rows are selected.
        action = enable if rows else disable
        action.append(self.delete_entry_bn)
        # Enable Save Diary there are entries.
        action = enable if has_entries else disable
        action.append(self.save_diary_bn)
        # Enable Delete Metric if a metric is selected.
        action = enable if self.metric_field.currentIndex() != 0 else disable
        action.append(self.delete_metric_bn)
        # Enable and disable buttons.
        for button in enable:
            button.setEnabled(True)
        for button in disable:
            button.setEnabled(False)

    # -------------------- NUTRIENT GUIDE MENU -------------------- #
    def update_guide(self, effective_date=None):
        """
        Update the Nutrient Guide.

        Clears the Nutrient Guide Viewer table and repopulates it with the
        user's nutrient targets for the most recent effective date, or for the
        effective_date arg if it is not None. If no effective dates exist, the
        table will be empty. The table format is:

                                    'Value'         'Unit'
            First Nutrient Name     target value    target value
            Second Nutrient Name    target value    target value

        Each time the table is populated, the rows are automatically sorted by
        the nutrient names in the vertical header. Target values in this table
        are always for the nutrient's corresponding unit of measure. If the
        user identifies targets using the '% DV' unit (percent daily value),
        the converted target value will be shown.

        :param effective_date: Effective date for which targets are displayed,
            or None to display the most recent effective date's targets if one
            exists or, if not, a blank table
        :type effective_date: str, None
        """
        self.effective_field.blockSignals(True)
        self.guideviewer.setRowCount(0)
        self.guideviewer.setColumnCount(0)
        self.effective_field.clear()
        self.effective_field.addItem("-CHOOSE DATE-")
        # If the Nutrient Guide has targets, add them to the table.
        nutrient_guide = self.user.nutrient_guide()
        if not nutrient_guide:
            self.effective_tag.setText("NO TARGETS SET")
            self.effective_field.blockSignals(False)
            self.toggle_guide()
            return
        # Reverse sorted dates to list in descending order.
        reversed_effective_dates = sorted(nutrient_guide.keys())[::-1]
        self.guideviewer.setColumnCount(2)
        self.guideviewer.setRowCount(len(dna.GUI_NUTRIENTS))
        self.guideviewer.setHorizontalHeaderLabels(["Value", "Unit"])
        self.guideviewer.verticalHeader().setFixedWidth(190)
        self.guideviewer.setVerticalHeaderLabels(
            [dna.NUTRIENTS[nid][0] for nid in dna.GUI_NUTRIENTS])
        if effective_date is None:
            targets_date = reversed_effective_dates[0]
        else:
            targets_date = effective_date
        for nid in dna.GUI_NUTRIENTS:
            row = dna.GUI_NUTRIENTS.index(nid)
            value_item = QtGui.QTableWidgetItem()
            if nid in nutrient_guide[targets_date]:
                value_item.setText(str(nutrient_guide[targets_date][nid]))
                value_item.setTextAlignment(QtCore.Qt.AlignCenter)
            self.guideviewer.setItem(row, 0, value_item)
            unit_item = QtGui.QTableWidgetItem()
            unit_item.setText(str(dna.NUTRIENTS[nid][2]))
            unit_item.setTextAlignment(QtCore.Qt.AlignCenter)
            self.guideviewer.setItem(row, 1, unit_item)
        # Add all effective dates to Effective Date Field
        self.effective_field.addItems(reversed_effective_dates)
        self.effective_field.setCurrentIndex(
            self.effective_field.findText(targets_date))
        self.effective_tag.setText("Effective Date:  " + targets_date)
        self.effective_field.blockSignals(False)
        self.toggle_guide()

    def add_targets(self):
        """Add targets to the Nutrient Guide."""
        dialog = ears.InputTargets(self, self.user)
        if dialog.exec_():
            self.user.add_targets(dialog.targets)
            self.update_guide()
            self.toggle_plot()
            self.update_buildinfo(rebuild=True)

    def view_targets(self):
        """View targets for the selected date in the Nutrient Guide Viewer."""
        effective_date = str(self.effective_field.currentText())
        self.update_guide(effective_date)

    def delete_targets(self):
        """Delete targets for the selected date from the Nutrient Guide."""
        effective_date = str(self.effective_field.currentText())
        # Check that the user wants to delete the targets for this date.
        choice = QtGui.QMessageBox.warning(
            self, "Confirm Delete Targets", "Are you sure you want to " +
            "delete targets for the effective date:\n\n" + effective_date,
            QtGui.QMessageBox.Yes | QtGui.QMessageBox.No)
        if choice == QtGui.QMessageBox.No:
            return
        self.user.remove_targets(effective_date)
        self.update_guide()
        self.toggle_plot()
        self.update_buildinfo(rebuild=True)

    def toggle_guide(self):
        """
        Toggle the Nutrient Guide Menu.

        Enables or disables Nutrient Guide Menu buttons based on the user
        attribute and the Effective Date Field value.
        """
        date_idx = self.effective_field.currentIndex()
        enable = []
        disable = []
        # Enable Add Targets if user is logged in.
        action = enable if self.user is not None else disable
        action.append(self.add_targets_bn)
        # Enable View Targets if an effective date is selected.
        action = enable if date_idx != 0 else disable
        action.append(self.view_targets_bn)
        # Enable Delete Targets button if an effective date is selected.
        action = enable if date_idx != 0 else disable
        action.append(self.delete_targets_bn)
        # Enable and disable buttons.
        for button in enable:
            button.setEnabled(True)
        for button in disable:
            button.setEnabled(False)

    # --------------------- DATA PLOTTER MENU --------------------- #
    def plot_action(self, action_num):
        """
        Display the selected Data Plotter plot dialog.

        Calls the applicable plot dialog for the action_num arg. All plot
        dialog constructors are passed the user attribute. The performance
        results plot dialogs are also passed an Exercise property.

        :param action_num: Plot action number: 0-Health Diary, 1-Nutrient
            Targets, 2-Macronutrient Proportions, 3-Meal Times, 4-Nutrient
            Values, 5-Performance Results Per Exercise, 6-Performance Results
            Per Muscle, 7-Performance Results Per Tag, 8-Workout Periods
        :type action_num: int
        """
        plot_args = {
            0: "PlotHealthDiary",
            1: "PlotNutrientGuide",
            2: "PlotMacroProportions",
            3: "PlotMealTimes",
            4: "PlotNutrientValues",
            5: "itemid",
            6: "focusmuscle",
            7: "tags",
            8: "PlotWorkoutPeriods"
        }
        if action_num not in [5, 6, 7]:
            plot = getattr(eyes, plot_args[action_num])(self, self.user)
        else:
            plot = eyes.PlotPerformance(self, self.user, plot_args[action_num])
        if plot.exec_():
            pass

    def toggle_plot(self):
        """
        Toggle Data Plotter Menu.

        Enables or disables Data Plotter Menu buttons based on the user
        attribute and the sizes of the user's Health Diary, Nutrient Guide,
        Diet records, and Program records. Each user data source must have at
        least one data set to enable its corresponding menu button.
        """
        enable = []
        disable = []
        # Disable all buttons if a user is not logged in.
        if self.user is None:
            for idx in range(9):
                self.plot_bns[idx].setEnabled(False)
            return
        # Enable Health Diary button if it has entries.
        action = enable if self.user.health_diary() else disable
        action.append(self.plot_bns[0])
        # Enable Nutrient Guide button if it has targets.
        action = enable if self.user.nutrient_guide() else disable
        action.append(self.plot_bns[1])
        # Enable Diet Record inventory buttons if Diet records exist.
        action = enable if self.user.record_objects("D") else disable
        action.extend([self.plot_bns[idx] for idx in range(2, 5)])
        # Enable Program Record inventory buttons if Program records exist.
        action = enable if self.user.record_objects("P") else disable
        action.extend([self.plot_bns[idx] for idx in range(5, 9)])
        # Enable and disable buttons.
        for button in enable:
            button.setEnabled(True)
        for button in disable:
            button.setEnabled(False)

    # ------------------------- BUILD MENU ------------------------ #
    def build_action(self, action_num):
        """
        Execute the selected Build Menu action.

        Calls the applicable method for the action_num arg. Methods which alter
        the selected build element are passed that element, and those that do
        not are called without args. The move_element method is called with an
        additional arg indicating which direction the element is moved within
        the Build Viewer tree. The expand_children method can take a while to
        process, especially if the Build Parent is a fully filled Program, so a
        MessageDialog is instantiated to inform the user in case the UI hangs.

        :param action_num: Build action number: 0-Create New Build, 1-Move
            Element Down, 2-Move Element Up, 3-Add New Element, 4-Add Inventory
            Item, 5-Edit Element, 6-Remove Element, 7-Copy Element, 8-Paste
            Element, 9-Expand Children, 10-Collapse Children, 11-Save Element
            Template, 12-Save Build Record, 13-Clear Build Viewer
        :type action_num: int
        """
        current_element = self.buildviewer.currentItem()
        if action_num == 0:
            self.create_build()
        elif action_num == 1:
            self.move_element(current_element, -1)
        elif action_num == 2:
            self.move_element(current_element, 1)
        elif action_num == 3:
            self.add_element(current_element)
        elif action_num == 4:
            self.add_item(current_element)
        elif action_num == 5:
            self.edit_element(current_element)
        elif action_num == 6:
            self.remove_element(current_element)
        elif action_num == 7:
            self.copy_element(current_element)
        elif action_num == 8:
            self.paste_element(current_element)
        elif action_num == 9:
            self.expanding_message = organs.MessageDialog(
                self, "Expanding Children", "THIS MAY TAKE A WHILE...")
            self.expanding_message.show()
            QtGui.QApplication.processEvents()
            self.expand_children(current_element)
        elif action_num == 10:
            self.collapse_children(current_element)
        elif action_num == 11:
            self.save_template(current_element)
        elif action_num == 12:
            self.save_record()
        elif action_num == 13:
            self.clear_build()

    def create_build(self):
        """
        Create a new build.

        Opens an InputBuildElement dialog to let the user select a new build
        element to assign to the Build Parent. Calls InputBuildElement with the
        new build element to let the user edit its properties. If the user
        accepts both dialogs, the Build Parent is assigned the new build
        element, and the Build Viewer and Build Info Viewer are both updated.
        """
        bp_dialog = ears.InputBuildParent(self)
        if bp_dialog.exec_():
            new_buildelement = bp_dialog.build_parent
            be_dialog = ears.InputBuildElement(self, new_buildelement)
            if be_dialog.exec_():
                self.buildviewer.set_buildparent(new_buildelement)
                self.toggle_build()
                cid = self.buildviewer.buildparent().cid()
                idx = 1 if cid in dna.NUTRITION_BUILD_ELEMENTS else 2
                self.buildinfoviewer.setCurrentIndex(idx)
                self.update_buildinfo()

    def move_element(self, selected_element, direction):
        """
        Move the selected build element up or down in position.

        Reassigns the selected_element arg's index in its parent container one
        position to the left (direction=-1) or right (direction=1). This is
        shown as an up or down move, respectively, in the Build Viewer.

        :param selected_element: The selected build element
        :type selected_element: Quantity, Ingredient, Recipe, Meal, Session,
            Activity, Activity, Workout, Cycle
        :param direction: Direction to move build element: -1 for a left/up
            move and 1 for a down/right move
        :type direction: int
        """
        # Store whether the element is expanded or not for reference below.
        is_expanded = selected_element.isExpanded()
        parent = selected_element.parent()
        element_idx = parent.indexOfChild(selected_element)
        element_item = parent.takeChild(element_idx)
        # Update parent's children and container attributes. If the easier way
        # is used (widgetize), current expand/collapse config of children is
        # lost because everything is collapsed.
        parent.insertChild(element_idx + direction, element_item)
        parent.move_child(selected_element, direction)
        self.buildviewer.setCurrentItem(selected_element)
        if is_expanded:
            selected_element.setExpanded(True)

    def add_element(self, selected_element):
        """
        Add a new element as a child to the selected build element.

        Instantiates a new build element to add as a child to the selected
        build element. Calls InputBuildElement with the new child to let the
        user edit its properties. If the dialog is accepted, the selected
        build element's container and children attributes are updated with the
        new child. Updates Build Info Viewer.

        :param selected_element: The selected build element
        :type selected_element: Recipe, Meal, Diet, Workout, Cycle, Program
        """
        child_cid = dna.BUILD_ELEMENTS[selected_element.cid()][1][0]
        child_element = body.build_element(child_cid)
        dialog = ears.InputBuildElement(self, child_element)
        if dialog.exec_():
            selected_element.addChild(child_element)
            selected_element.add_child(child_element)
            self.buildviewer.setCurrentItem(child_element)
            self.update_buildinfo()

    def add_item(self, selected_element):
        """
        Add an inventory item as a child to the selected build element.

        Instantiates a new build element to add as a child to the selected
        build element. The child's state is set based on the selected inventory
        item. Calls the applicable input dialog with the new child to let the
        user edit its properties or children. If the user accepts the dialog,
        the selected build element's container and children attributes are
        updated with the new child. Updates Build Info Viewer.

        :param selected_element: The selected build element
        :type selected_element: Recipe, Meal, Diet, Workout, Cycle, Program
        """
        inventoryid = self.inventory_field.currentIndex()
        selected_inv_row = self.inventoryviewer.currentRow()
        itemid = str(self.inventoryviewer.item(selected_inv_row, 0).text())
        cid = INVENTORY_CIDS[inventoryid]
        if inventoryid in [0, 1]:
            item_element = body.build_element(cid, itemid)
            class_name = "Input" + dna.BUILD_ELEMENTS[cid][0]
            dialog = getattr(ears, class_name)(self, item_element)
        else:
            item_element = body.build_element(cid)
            item_element.set_state(self.user.template_state(cid, itemid))
            item_element.widgetize()
            dialog = ears.InputBuildElement(self, item_element)
        if dialog.exec_():
            selected_element.addChild(item_element)
            selected_element.add_child(item_element)
            # Child item not selected-better behavior when adding multiple
            # Ingredients or Activities to the same build element. Toggle
            # Build Menu.
            self.buildviewer.expandItem(selected_element)
            self.buildviewer.expandItem(item_element)
            self.toggle_build()
            self.update_buildinfo()

    def edit_element(self, selected_element):
        """
        Edit the selected build element.

        Calls the applicable input dialog with the selected build element to
        let the user edit its properties or children. If the user accepts the
        dialog, the selected build element's properties/children are directly
        updated by the dialog. Updates Build Info Viewer.

        :param selected_element: The selected build element
        :type selected_element: Ingredient, Recipe, Meal, Diet, Activity,
            Workout, Cycle, Program
        """
        cid = selected_element.cid()
        if cid == "I":
            dialog = ears.InputIngredient(self, selected_element)
        elif cid == "A":
            dialog = ears.InputActivity(self, selected_element)
        else:
            dialog = ears.InputBuildElement(self, selected_element)
        if dialog.exec_():
            if cid in ["I", "A"]:
                selected_element.setExpanded(True)
            # Toggle buttons to enable expand/collapse buttons when adding
            # new Quantity or Session objects.
            self.toggle_build()
            self.update_buildinfo()

    def remove_element(self, selected_element):
        """
        Remove the selected build element.

        Removes the selected build element from its parent's children and
        container attributes. Updates Build Info Viewer.

        :param selected_element: The selected build element
        :type selected_element: Quantity, Ingredient, Recipe, Meal, Session,
            Activity, Workout, Cycle
        """
        parent = selected_element.parent()
        parent.removeChild(selected_element)
        parent.remove_child(selected_element)
        # Toggle buttons - ensures that expand/collapse get toggled when
        # deleting all children. In this case, parent is selected first and
        # buttons are toggled before children are actually deleted.
        self.toggle_build()
        self.update_buildinfo()

    def copy_element(self, selected_element):
        """
        Copy the selected build element.

        Copies the selected build element, displays its description in the
        Copied Viewer, and assigns the copy to the copied_element attribute.

        :param selected_element: The selected build element
        :type selected_element: Ingredient, Recipe, Meal, Activity, Workout,
            Cycle
        """
        cid = selected_element.cid()
        itemid = selected_element.itemid() if cid in ["I", "A"] else None
        self.copied_element = body.build_element(cid, itemid)
        self.copied_element.set_state(selected_element.state())
        self.copied_element.widgetize()
        self.copiedviewer.setText(str(self.copied_element))
        # Toggle Build Menu in case copied element is a Recipe, which can be
        # pasted to itself.
        self.toggle_build()

    def paste_element(self, selected_element):
        """
        Add the copied element as a child to the selected build element.

        Updates the selected build element's container and children attributes
        with the copied element. Copied elements can only be pasted once, so
        the copied_element attribute is assigned None and the Copied Viewer
        text is cleared. Updates Build Info Viewer.

        :param selected_element: The selected build element
        :type selected_element: Recipe, Meal, Diet, Workout, Cycle, Program
        """
        selected_element.addChild(self.copied_element)
        selected_element.add_child(self.copied_element)
        self.buildviewer.setCurrentItem(self.copied_element)
        self.buildviewer.expandItem(self.copied_element)
        self.copied_element = None
        self.copiedviewer.clear()
        self.update_buildinfo()

    def expand_children(self, element, depth=0):
        """
        Expand all children under the selected build element.

        Recursively calls itself while incrementing the depth arg by 1 until
        all constituent build element children of the selected build element
        are expanded. Terminates when there are no more children to expand, or
        the element arg is a Quantity or Session--which both return zero when
        their inherited childCount method is called. When the depth arg is zero
        and all children have been expanded, the MessageDialog object assigned
        to the expanding_message attribute by the caller is closed, and None is
        then assigned to the attribute.

        :param element: The selected build element or one of its children
        :type element: Quantity, Ingredient, Recipe, Meal, Diet, Session,
            Activity, Workout, Cycle, Program
        :param depth: Depth at which the recursive call to this method is made,
            with the top call == 0 and all lower calls having depths that have
            been incremented by +1 for each level down
        :type depth: int
        """
        # If the element has an empty children attribute, return.
        if depth > 0 and element.childCount() == 0:
            return
        # Expand item and its children recursively.
        self.buildviewer.expandItem(element)
        for child_idx in range(element.childCount()):
            self.expand_children(element.child(child_idx), depth + 1)
        # Close the expanding message created by change_build.
        if depth == 0:
            self.expanding_message.finish()
            self.expanding_message = None

    def collapse_children(self, element):
        """
        Collapse all children under the selected build element.

        Recursively calls itself until all constituent build element children
        of the selected build element are collapsed. Terminates when there are
        no more children to expand, or the element arg is a Quantity or Session
        --which both return zero when their inherited childCount method is
        called.

        :param element: The selected build element or one of its children
        :type element: Quantity, Ingredient, Recipe, Meal, Diet, Session,
            Activity, Workout, Cycle, Program
        """
        # If the element has an empty children attribute, return.
        if element.childCount() == 0:
            return
        # Collapse item and its children recursively.
        self.buildviewer.collapseItem(element)
        for child_idx in range(element.childCount()):
            self.collapse_children(element.child(child_idx))

    def save_template(self, selected_element):
        """
        Save the selected build element as a template.

        Checks the template description for the default value '(unnamed)' and,
        if found, verifies that the user wants to save anyway. Calls the user's
        add_template method with the selected_element's template state and
        verifies that the next template item ID has been returned. The user is
        notified of the result. Reloads the current inventory if necessary.

        :param selected_element: The selected build element
        :type selected_element: Recipe, Meal, Diet, Workout, Cycle, Program
        """
        cid = selected_element.cid()
        name = dna.BUILD_ELEMENTS[cid][0]
        # Check for no description given.
        if selected_element.description() == "(unnamed)":
            choice = QtGui.QMessageBox.warning(
                self, "Confirm Missing Description", "Are you sure you " +
                "want to save this template without a description? A " +
                "description makes it easier to search for using the " +
                "Inventory Manager.",
                QtGui.QMessageBox.Yes | QtGui.QMessageBox.No)
            if choice == QtGui.QMessageBox.No:
                return
        # Save the template and check for maximum template IDs reached.
        tid = self.user.add_template(selected_element.template_state())
        if not tid:
            QtGui.QMessageBox.warning(
                self, "Maximum Templates Reached",
                "You cannot save anymore " + name + " templates!")
            return
        # Reload template inventory if necessary.
        inventoryid = self.inventory_field.currentIndex()
        selectedrow = self.inventoryviewer.currentRow()
        if inventoryid in range(2, 8) and cid == INVENTORY_CIDS[inventoryid]:
            self.load_inventory(
                inventoryid, keep_order=True, selected_row=selectedrow)
        QtGui.QMessageBox.information(
            self, "Template Saved", "This " + name + " template has been " +
            "saved:\n\n" + tid + ": " + selected_element.description())

    def save_record(self):
        """
        Save the Diet or Program Build Parent as a record.

        Checks the Diet or Program build for valid data. If it is a Diet,
        checks its container for duplicate Meal times and, if found, verifies
        that the user wants to save anyway. If it is a Program, verifies that
        all constituent Workout periods begin on or after the start date, and
        that there are no duplicate Workout period began datetimes. In either
        case, the user cannot save the record. Checks if the record date exists
        in the applicable inventory and, if so, verifies that the user wants to
        overwrite the old record with the new one. The user is notified of the
        result. Reloads the current inventory if necessary.
        """
        buildparent = self.buildviewer.buildparent()
        cid = buildparent.cid()
        build_name = dna.BUILD_ELEMENTS[cid][0]
        build_state = buildparent.state()
        build_desc = build_state[2]
        build_date = build_state[3]
        if cid == "D":
            # Check for duplicate Meal times, confirm use wants to save anyway.
            meal_times = buildparent.meal_times()
            if len(meal_times) != len(set(meal_times)):
                choice = QtGui.QMessageBox.warning(
                    self, "Confirm Meal Times", "At least one Meal time has " +
                    "been duplicated! This will affect the Meal Times plot. " +
                    "Do you want to save this record anyway?",
                    QtGui.QMessageBox.Yes | QtGui.QMessageBox.No)
                if choice == QtGui.QMessageBox.No:
                    return
        if cid == "P":
            begans = []
            for cycle in buildparent.container():
                for workout in cycle.container():
                    begans.append(workout.began())
            # Check for Workout began date occurring prior to Program start.
            for began in begans:
                if str(began.date()) < build_date:
                    QtGui.QMessageBox.warning(
                        self, "Invalid Workout Period", "At least one " +
                        "Workout period begins before the Program start " +
                        "date. This Program record cannot be saved until " +
                        "all Workouts begin on or after:\n\n" + build_date)
                    return
            # Check for duplicate Workout period begans.
            if len(begans) != len(set(begans)):
                QtGui.QMessageBox.warning(
                    self, "Repeated Workout Period", "At least two Workout " +
                    "periods begin at the same date and time! If two or " +
                    "more Workouts with the same period-began date and time " +
                    "have Activity Sessions added to them, your Build Info " +
                    "and performance data plots will show inaccurate " +
                    "results.\n\nIf your record is based on a template and " +
                    "some Workouts have yet to be updated with future " +
                    "Activity Sessions and periods, you can ignore this " +
                    "warning. All Workouts in a template default to the " +
                    "date on which you first saved the template as a record.")
        # Check if the record exists, confirm overwrite record.
        if self.user.record_exists(build_state):
            date_string = "date" if cid == "D" else "start date"
            existing_object = self.user.record_objects(cid, build_date)
            desc = existing_object.description()
            choice = QtGui.QMessageBox.warning(
                self, "Confirm Overwrite Record", "A " + build_name + " " +
                "record already exists for this " + date_string + ". " +
                "Do you want to replace this record:\n\n" + build_date +
                ": " + desc, QtGui.QMessageBox.Yes | QtGui.QMessageBox.No)
            if choice == QtGui.QMessageBox.No:
                return
        # Save the record to the records inventory and toggle plot buttons.
        self.user.add_record(build_state)
        self.toggle_plot()
        # Reload inventory if necessary.
        inventoryid = self.inventory_field.currentIndex()
        selectedrow = self.inventoryviewer.currentRow()
        if inventoryid in [8, 9] and cid == INVENTORY_CIDS[inventoryid]:
            self.load_inventory(
                inventoryid, keep_order=True, selected_row=selectedrow)
        QtGui.QMessageBox.information(
            self, "Record Saved", "This " + build_name + " record has been " +
            "saved:\n\n" + build_date + ": " + buildparent.description())

    def clear_build(self):
        """
        Clear the Build Parent from the Build Viewer.

        Verifies that the user wants to clear the Build Viewer and, if so,
        clears the Build Parent from the Build Viewer and updates the Build
        Info Viewer.
        """
        choice = QtGui.QMessageBox.question(
            self, "Confirm Clear Build", "Are you sure you want to clear " +
            "your build?", QtGui.QMessageBox.Yes | QtGui.QMessageBox.No)
        if choice == QtGui.QMessageBox.No:
            return
        self.buildviewer.clear_buildparent()
        self.toggle_build()
        self.toggle_inventory()
        self.buildinfoviewer.setCurrentIndex(0)

    def toggle_build(self):
        """
        Toggle Build Menu.

        Enables or disables Build Menu buttons based on the Build Parent,
        the selected build element, the selected inventory item, the
        copied_element attribute, and the user's 'buildonly' setting.
        """
        # Disable all but Create New Build if no Build Parent, otherwise
        # disable Create New Build and assess conditions for remaining buttons.
        buildparent = self.buildviewer.buildparent()
        if buildparent is None:
            for button in self.build_bns[1:]:
                button.setEnabled(False)
            self.build_bns[0].setEnabled(True)
            return
        else:
            self.build_bns[0].setEnabled(False)
        enable = []
        disable = []
        selected_element = self.buildviewer.currentItem()
        cid = selected_element.cid()
        parent = selected_element.parent()
        # Enable Move Element Up/Down based on element parent/position.
        if parent is not None:
            index = parent.indexOfChild(selected_element)
            action = enable if index != 0 else disable
            action.append(self.build_bns[1])
            action = enable if index + 1 < parent.childCount() else disable
            action.append(self.build_bns[2])
        else:
            disable.extend([self.build_bns[idx] for idx in [1, 2]])
        # Enable Add New Element for specific elements.
        action = enable if cid in ["R", "M", "D", "C", "P"] else disable
        action.append(self.build_bns[3])
        # Enable Add Inventory Item if selected inventory item is a
        # child of the selected build element.
        inventoryid = self.inventory_field.currentIndex()
        selected_row = self.inventoryviewer.currentRow()
        if (selected_row != -1 and
           INVENTORY_CIDS[inventoryid] in dna.BUILD_ELEMENTS[cid][1]):
            action = enable
        else:
            action = disable
        action.append(self.build_bns[4])
        # Enable Edit Element for all build elements but Quantity/Session.
        action = enable if cid not in ["Q", "S"] else disable
        action.append(self.build_bns[5])
        # Enable Remove Element for all but Build Parent.
        action = enable if parent is not None else disable
        action.append(self.build_bns[6])
        # Enable Copy Element if not Quantity/Session and has a parent--unless
        # it is a Recipe (since Recipes can be added to Recipes).
        if cid not in ["Q", "S"] and (parent is not None or cid == "R"):
            action = enable
        else:
            action = disable
        action.append(self.build_bns[7])
        # Enable Paste Element if copied element exists and it is a child of
        # the selected build element.
        if (self.copied_element is not None and
           self.copied_element.cid() in dna.BUILD_ELEMENTS[cid][1]):
            action = enable
        else:
            action = disable
        action.append(self.build_bns[8])
        # Enable Expand/Collapse Children if selected element has children.
        action = enable if selected_element.childCount() > 0 else disable
        action.extend([self.build_bns[9], self.build_bns[10]])
        # Enable Save Element Template for the applicable class IDs.
        action = enable if cid in ["R", "M", "D", "W", "C", "P"] else disable
        action.append(self.build_bns[11])
        # Enable Save Build Record for the applicable class IDs.
        action = enable if buildparent.cid() in ["D", "P"] else disable
        action.append(self.build_bns[12])
        # Enable Clear Build Viewer if Build Parent is loaded.
        action = enable if buildparent is not None else disable
        action.append(self.build_bns[13])
        # Enable and disable buttons.
        for button in enable:
            button.setEnabled(True)
        for button in disable:
            button.setEnabled(False)
        # Update Build Info Viewer if applicable.
        if not self.user.settings()["buildonly"]:
            self.update_buildinfo()

    # ------------------------- BUILD INFO ------------------------ #
    def update_buildinfo(self, rebuild=False):
        """
        Update the Build Info Viewer.

        If the rebuild arg is False, clears the nutritioninfo and fitnessinfo
        attributes of all columns and rebuilds each. The new columns headers
        are the elements in the user's 'nutrients' and 'muscles' settings,
        respectively. Each table is populated with blank QTableWidgetItems. If
        the Build Parent is not None, calls the update_nutritioninfo or
        update_fitnessinfo method to fill the table with updated data. Each
        method is passed the applicable user setting with nutrients or muscles.

        :param rebuild: True if the table's structure is reconfigured due to a
            change in user, user settings, or nutrient targets, otherwise False
        :return: bool
        """
        user_nutrients = self.user.settings()["nutrients"]
        user_muscles = self.user.settings()["muscles"]
        if rebuild:
            # Restructure Nutrition Info.
            self.nutritioninfo.setColumnCount(0)
            self.nutritioninfo.setColumnCount(len(user_nutrients))
            self.nutritioninfo.setHorizontalHeaderLabels(
                [dna.NUTRIENTS[nid][1] for nid in user_nutrients])
            for col in range(len(user_nutrients)):
                for row in range(3):
                    item = QtGui.QTableWidgetItem()
                    item.setTextAlignment(QtCore.Qt.AlignCenter)
                    self.nutritioninfo.setItem(row, col, item)
                    if row == 1:
                        self.nutritioninfo.item(row, col).setText(
                            dna.NUTRIENTS[user_nutrients[row]][2])
            # Restructure Fitness Info.
            self.fitnessinfo.setColumnCount(0)
            self.fitnessinfo.setColumnCount(len(user_muscles))
            self.fitnessinfo.setHorizontalHeaderLabels(user_muscles)
            for col in range(len(user_muscles)):
                for row in range(3):
                    item = QtGui.QTableWidgetItem()
                    item.setTextAlignment(QtCore.Qt.AlignCenter)
                    self.fitnessinfo.setItem(row, col, item)
        # If Build Parent exists, update applicable info table.
        buildparent = self.buildviewer.buildparent()
        if buildparent is not None:
            if buildparent.cid() in dna.NUTRITION_BUILD_ELEMENTS:
                self.update_nutritioninfo(user_nutrients)
            else:
                self.update_fitnessinfo(user_muscles)

    def update_nutritioninfo(self, user_nutrients):
        """
        Update the Build Info Viewer's Nutrition Info table.

        Populates the nutritioninfo attribute's cells with nutrient data for
        the Build Parent or, if the user's 'buildonly' setting is False, the
        selected build element. Calculates the value of each nutrient in the
        user_nutrients arg and places it in the corresponding column of the
        first row. Finds the latest targets in the Nutrient Guide, calculates
        each nutrient's percent-of-target value for the third row's data, and
        sets the background color of each cell in that row based on the
        percent. If no targets are found, all values in that row are '0'.

        :param user_nutrients: The user's 'nutrients' setting
        :type user_nutrients: list
        """
        nutrient_guide = self.user.nutrient_guide()
        latest_date = organs.max_key(nutrient_guide)
        targs = {} if latest_date is None else nutrient_guide[latest_date]
        # Determine if nutrient values are for build or selected build element.
        buildonly = self.user.settings()["buildonly"]
        if buildonly:
            element = self.buildviewer.buildparent()
        else:
            element = self.buildviewer.currentItem()
        # Calculate applicable nutrient values and populate table.
        nvals = {nid: element.nutrient_value(nid) for nid in user_nutrients}
        for nid in user_nutrients:
            col = user_nutrients.index(nid)
            # Place the nutrient value in first row.
            value = str(organs.numeric_value(nvals[nid], round_digits=2))
            self.nutritioninfo.item(0, col).setText(value)
            # Place the target percent in the third row.
            if nid in targs:
                pct = "{0:.1f}".format((nvals[nid] / float(targs[nid])) * 100)
            else:
                pct = "0"
            self.nutritioninfo.item(2, col).setText(pct)
            # Change background colors based on blank text or numeric value.
            rgbs = {
                100: [255, 185, 173], 80: [255, 222, 173],
                60: [241, 255, 173], 40: [200, 255, 173],
                20: [173, 255, 230], 0: [246, 229, 255]}
            rgb = None
            for rgb_pct in sorted(rgbs.keys())[::-1]:
                if float(pct) > rgb_pct:
                    rgb = rgbs[rgb_pct]
                    break
            if rgb is None:
                rgb = [227, 229, 229]
            self.nutritioninfo.item(2, col).setBackground(QtGui.QColor(*rgb))
            # Set header tool tip.
            self.nutritioninfo.horizontalHeaderItem(col).setToolTip(
                dna.NUTRIENTS[nid][0])
        self.nutritioninfo.resizeColumnsToContents()

    def update_fitnessinfo(self, user_muscles):
        """
        Update the Build Info Viewer's Fitness Info table.

        Populates the fitnessinfo attribute's cells with muscle data for the
        Build Parent or, if the user's 'buildonly' setting is False, the
        selected build element. Calculates the number of sessions, total
        effort, and maximum intensity for each muscle in the user_muscles arg
        and places it in the corresponding cell. Sets the background color of
        each cell based on whether or not it shows a value for a muscle.

        :param user_muscles: The user's 'muscles' setting
        :type user_muscles: list
        """
        # Determine if muscle values are for build or selected build element.
        buildonly = self.user.settings()["buildonly"]
        if buildonly:
            element = self.buildviewer.buildparent()
        else:
            element = self.buildviewer.currentItem()
        # Calculate muscle data and populate table.
        sessions = element.muscle_sessions()
        effs = element.performance_results("effort", "focusmuscle")
        ints = element.performance_results("intensity", "focusmuscle")
        # Aggregate effort and intensity values for all Workouts.
        agg_effs = {}
        agg_ints = {}
        if element.cid() in ["S", "A"]:
            agg_effs = effs
            agg_ints = ints
        else:
            for wo_began in effs:
                agg_effs = organs.summed_dicts(agg_effs, effs[wo_began])
                agg_ints = organs.maxed_dicts(agg_ints, ints[wo_began])
        # Change blank cell colors to gray and filled cells to white.
        white, gray = [255, 255, 255], [227, 229, 229]
        for col in range(self.fitnessinfo.columnCount()):
            col_mus = user_muscles[col]
            ses_val = "" if col_mus not in sessions else str(sessions[col_mus])
            eff_val = "" if col_mus not in agg_effs else str(agg_effs[col_mus])
            int_val = "" if col_mus not in agg_ints else str(agg_ints[col_mus])
            self.fitnessinfo.item(0, col).setText(ses_val)
            self.fitnessinfo.item(1, col).setText(eff_val)
            self.fitnessinfo.item(2, col).setText(int_val)
            ses_rgb = gray if ses_val == "" else white
            eff_rgb = gray if eff_val == "" else white
            int_rgb = gray if int_val == "" else white
            self.fitnessinfo.item(0, col).setBackground(QtGui.QColor(*ses_rgb))
            self.fitnessinfo.item(1, col).setBackground(QtGui.QColor(*eff_rgb))
            self.fitnessinfo.item(2, col).setBackground(QtGui.QColor(*int_rgb))
        self.fitnessinfo.resizeColumnsToContents()

    # ---------------------- INVENTORY MENU ----------------------- #
    def load_inventory(self, inv_id, keep_order=False, selected_row=None):
        """
        Load the selected inventory into the Inventory Viewer.

        Determines the inventoryviewer attribute's headers and items based on
        the inventory type selected in the Inventory Field. Builds the table
        with applicable headers, item IDs, item descriptions, and, if the
        Exercise reference inventory is selected, performance metric units.
        Inventory items are filtered to show only favorites if the
        view_favorites attribute is True, i.e. the View Favorite Items button
        in the Inventory Menu is checked.

        If the keep_order arg is True and the selected_row arg is not None,
        the inventory is being reloaded after a change to its items. The item
        order and currently selected item are maintained for the user's
        convenience. The table is updated with added/removed items if
        applicable.

        :param inv_id: The inventory ID of the inventory selected in the
            Inventory Field--passed explicitly by callers or via QComboBox
            signal
        :type inv_id: int
        :param keep_order: True to keep the inventory item order, or False to
            rebuild the entire table and scroll to the top
        :type keep_order: bool
        :param selected_row: The currently selected row in the pre-reloaded
            inventory that will be reselected after it has been reloaded
        :type selected_row: int
        """
        cid = INVENTORY_CIDS[inv_id]
        # Inventory type headers, item getters, item info getters.
        parts = {
            0: (["Description"], "food_details"),
            1: (["Description", "Units"], "exercise_details")}
        for idx in range(2, 8):
            parts[idx] = (["Description"], "templates", "template_state")
        for idx in range(8, 10):
            parts[idx] = (["Description"], "records", "record_state")
        # Determine item IDs.
        if self.view_favorites:
            faves = self.user.settings()["favorites"]
            item_ids = faves[INVENTORY_FAVE_IDS[inv_id]]
        elif inv_id in [0, 1]:
            item_ids = getattr(self.user, parts[inv_id][1])().keys()
        else:
            item_ids = getattr(self.user, parts[inv_id][1])(cid).keys()
        # Determine item info and pair with item IDs.
        item_info = {}
        if inv_id in [0, 1]:
            for itemid in item_ids:
                dets = getattr(self.user, parts[inv_id][1])(itemid)
                info = [dets[0]] if inv_id != 1 else [dets[0], dets[2]]
                item_info[itemid] = info
        else:
            for itemid in item_ids:
                item_info[itemid] = [
                    getattr(self.user, parts[inv_id][2])(cid, itemid)[2]]
        # Determine final IDs and info based on parameters.
        sort_ord = 1 if self.user.settings()["sortup"] else -1
        sort_idx = 0 if self.user.settings()["sortid"] else 1
        if not item_ids:
            # Item info is not needed if no final IDs.
            final_ids = []
        elif keep_order:
            row_ids = []
            for row in range(self.inventoryviewer.rowCount()):
                row_ids.append(str(self.inventoryviewer.item(row, 0).text()))
            # Deal with items recently added or removed from the table.
            rem_items = list(set(row_ids) - set(item_info.keys()))
            for itemid in rem_items:
                row_ids.remove(itemid)
            add_items = list(set(item_info.keys()) - set(row_ids))
            final_ids = row_ids + add_items
            final_des = [item_info[itemid] for itemid in final_ids]
        else:
            final_ids, final_des = zip(*sorted(
                item_info.items(), key=lambda x: x[sort_idx])[::sort_ord])
        # Update inventory tag.
        inventorytag = INVENTORY_NAMES[inv_id]
        inventorytag += " (FAVORITES)" if self.view_favorites else ""
        self.inventory_tag.setText(inventorytag)
        # Reset table with applicable parts, item IDs, and item info.
        num_cols = 2 if inv_id != 1 else 3
        headers = ["Item ID"] + parts[inv_id][0]
        self.inventoryviewer.blockSignals(True)
        self.inventoryviewer.setRowCount(0)
        self.inventoryviewer.setColumnCount(0)
        self.inventoryviewer.setRowCount(len(final_ids))
        self.inventoryviewer.setColumnCount(num_cols)
        self.inventoryviewer.setHorizontalHeaderLabels(headers)
        for row in range(len(final_ids)):
            id_item = QtGui.QTableWidgetItem()
            id_item.setText(final_ids[row])
            id_item.setTextAlignment(QtCore.Qt.AlignCenter)
            id_item.setFont(FONT_TAGD)
            self.inventoryviewer.setItem(row, 0, id_item)
            desc_item = QtGui.QTableWidgetItem()
            desc_item.setText(final_des[row][0])
            self.inventoryviewer.setItem(row, 1, desc_item)
            if inv_id == 1:
                met_item = QtGui.QTableWidgetItem()
                met_item.setText(
                    final_des[row][1][0] + " -- " + final_des[row][1][1])
                met_item.setTextAlignment(QtCore.Qt.AlignCenter)
                self.inventoryviewer.setItem(row, 2, met_item)
        # Finalize column widths.
        self.inventoryviewer.resizeColumnsToContents()
        if inv_id == 1:
            self.inventoryviewer.horizontalHeader().setResizeMode(
                QtGui.QHeaderView.Interactive)
            self.inventoryviewer.setColumnWidth(1, 350)
        self.inventoryviewer.horizontalHeader().setStretchLastSection(True)
        # Process events to so table is built before row selection. Re-enable
        # signals and toggle Inventory/Build Menus.
        QtGui.QApplication.processEvents()
        if selected_row is not None:
            choose_row = selected_row
            row_count = self.inventoryviewer.rowCount()
            if selected_row == -1 or row_count == 0:
                self.inventoryviewer.blockSignals(False)
                self.toggle_inventory()
                self.toggle_build()
                return
            if selected_row > row_count - 1:
                choose_row = row_count - 1
            self.inventoryviewer.scrollToItem(
                self.inventoryviewer.item(choose_row, 0),
                QtGui.QAbstractItemView.PositionAtCenter)
            self.inventoryviewer.selectRow(choose_row)
        self.inventoryviewer.blockSignals(False)
        self.toggle_inventory()
        self.toggle_build()

    def search_inventory(self):
        """
        Search inventory for items whose descriptions contain all search terms.

        Parses search terms and finds all items with descriptions that contain
        all space-separated search terms. Hides all rows containing items with
        descriptions that do not match all terms. If no items are found, all
        rows are hidden. Searches are not case-sensitive.
        """
        search_terms = str(self.search_field.text()).strip()
        if search_terms == "":
            return
        # Unhide all rows.
        for row in range(self.inventoryviewer.rowCount()):
            self.inventoryviewer.setRowHidden(row, False)
        # Split search terms and determine all matched rows.
        words = search_terms.split(" ")
        show_rows = set(range(self.inventoryviewer.rowCount()))
        for word in words:
            match_rows = set([
                item.row() for item in self.inventoryviewer.findItems(
                    word, QtCore.Qt.MatchContains)])
            show_rows = show_rows & match_rows
        # Hide rows without matches.
        for row in range(self.inventoryviewer.rowCount()):
            if row not in show_rows:
                self.inventoryviewer.setRowHidden(row, True)

    def toggle_search(self, search_field_text):
        """
        Toggle the Search Inventory button and unhide inventory items.

        Enables or disables the Search Inventory button based on the
        search_field_text arg, passed by the Search Field's textChanged signal.
        If the arg is an empty QString, i.e. the Search Field is empty, the
        Search button is disabled and all rows in the inventoryviewer attribute
        are unhidden. If the arg is a non-empty QString, the Search button is
        enabled.

        :param search_field_text: Current text in the Search Field, emitted
            by its textChanged signal, or text passed to this method by a
            caller, or None
        :param search_field_text: QString
        """
        if search_field_text == "":
            for row in range(self.inventoryviewer.rowCount()):
                self.inventoryviewer.setRowHidden(row, False)
            self.search_bn.setEnabled(False)
        else:
            self.search_bn.setEnabled(True)

    def inventory_action(self, action_num):
        """
        Execute the selected Inventory Menu action.

        Calls the applicable method for the action_num arg. All methods are
        passed the index of the inventory that is currently selected in the
        Inventory Field--the 'inventory ID'. Methods which access the inventory
        item are passed its item ID. Methods which alter the currently loaded
        inventory are passed the selected row so that it may be reselected
        after the inventory is updated. The view_item method is passed no args
        since it is connected to the inventoryviewer's itemDoubleClicked signal
        which carries no args.

        :param action_num: Action number, 0-Send Item to Build, 1-Create New
            Reference, 2-Load Data Capsules, 3-View Item, 4-Edit Item, 5-Save
            Data Capsule, 6-Delete Item, 7-View All Items, 8-View Favorite
            Items
        :type action_num: int
        """
        inventoryid = self.inventory_field.currentIndex()
        # Item IDs will be irrelevant if no item is selected.
        selectedrow = self.inventoryviewer.currentRow()
        if selectedrow != -1:
            itemid = str(self.inventoryviewer.item(selectedrow, 0).text())
        if action_num == 0:
            self.send_item(inventoryid, itemid)
        elif action_num == 1:
            self.create_reference(inventoryid, selectedrow)
        elif action_num == 2:
            self.load_data_capsule(inventoryid, selectedrow)
        elif action_num == 3:
            self.view_item()
        elif action_num == 4:
            self.edit_item(inventoryid, itemid, selectedrow)
        elif action_num == 5:
            self.save_data_capsule(inventoryid, itemid)
        elif action_num == 6:
            self.delete_item(inventoryid, itemid, selectedrow)
        elif action_num == 7:
            self.view_favorites = False
            self.load_inventory(inventoryid)
        elif action_num == 8:
            self.view_favorites = True
            self.load_inventory(inventoryid)

    def send_item(self, inventoryid, itemid):
        """
        Send the selected item to the Build Viewer as the Build Parent.

        Initializes a build element object with the same state as the inventory
        item and widgetizes it to display it accurately in the Build Viewer.
        The build element is passed to the buildviewer attribute as the Build
        Parent and the Build Info Viewer is updated.

        :param inventoryid: The index of the currently selected inventory
        :type inventoryid: int
        :param itemid: The item ID of the selected inventory item
        :type itemid: str
        """
        cid = INVENTORY_CIDS[inventoryid]
        item_element = body.build_element(cid)
        if inventoryid in range(2, 8):
            item_element.set_state(self.user.template_state(cid, itemid))
        else:
            item_element.set_state(self.user.record_state(cid, itemid))
        item_element.widgetize()
        # Set Build Parent and toggle Build Menu.
        self.buildviewer.set_buildparent(item_element)
        self.buildviewer.expandItem(item_element)
        self.toggle_build()
        # Update Build Info and toggle Inventory Menu.
        self.update_buildinfo()
        index = 1 if cid in dna.NUTRITION_BUILD_ELEMENTS else 2
        self.buildinfoviewer.setCurrentIndex(index)
        self.toggle_inventory()

    def create_reference(self, inventoryid, selectedrow):
        """
        Create a Food or Exercise reference item.

        Opens the applicable input dialog and lets the user edit the new
        reference item's properties. If the dialog is accepted, calls the
        user's add_reference method with the dialog's refobject attribute and
        verifies that the next reference ID has been returned. The user is
        notified of the result. Reloads the current inventory.

        :param inventoryid: The index of the currently selected inventory
        :type inventoryid: int
        :param selectedrow: The selected row in the Inventory Viewer table
        :type selectedrow: int
        """
        reftype = "Food" if inventoryid == 0 else "Exercise"
        dialog = getattr(ears, "Input" + reftype)(self, self.user)
        if dialog.exec_():
            # Save the reference and check for maximum reference IDs reached.
            ref = getattr(dialog, reftype.lower())
            rid = self.user.add_reference(ref)
            if not rid:
                QtGui.QMessageBox.warning(
                    self, "Maximum References Reached",
                    "You cannot save anymore " + reftype + " items!")
                return
            # Reload inventory.
            self.load_inventory(
                inventoryid, keep_order=True, selected_row=selectedrow)
            QtGui.QMessageBox.information(
                self, "Reference Item Created", "This " + reftype + " item " +
                "has been created:\n\n" + rid + ": " + ref.info("description"))

    def load_data_capsule(self, inventoryid, selectedrow):
        """
        Creates one or more reference items from data capsules.

        Opens a DataCapsuleDialog to let the user select one or more data
        capsule files, with the .json extension, from which to create new
        reference items. A valid data capsule has the following data structure:

        --Food Data Capsule--
            ['Food', description, groupid, [unitsequences], {nutrientcontent}]
        --Exercise Data Capsule--
            ['Exercise', description, focusmuscle, [units], [tags]]

        For each selected file, its data are passed to a Reference constructor
        and the Reference is passed to the applicable input dialog that let's
        the user edit its properties before saving. Thorough validation of the
        data capsule's contents occur during Reference instantiation. If any
        data are found to be invalid, an exception is raised and the user is
        informed that a reference item cannot be created from the file. If the
        user accepts the input dialog, the user's add_reference method is
        called with the dialog's refobject attribute. Each time a new reference
        is successfully added to the inventory, the user is notified and the
        inventory is reloaded.

        :param inventoryid: The index of the currently selected inventory
        :type inventoryid: int
        :param selectedrow: The selected row in the Inventory Viewer table
        :type selectedrow: int
        """
        file_dialog = organs.EasyFileDialog(self, "loadcapsule")
        if not file_dialog.exec_():
            return
        q_paths = file_dialog.selectedFiles()
        capsule_paths = [os.path.normpath(str(q_path)) for q_path in q_paths]
        # Check for no files selected.
        if not capsule_paths:
            return
        reftype = "Food" if inventoryid == 0 else "Exercise"
        num_caps = len(capsule_paths)
        # If more than one file was selected, inform the user that a series of
        # input dialogs, one for each, will be presented.
        if num_caps > 1:
            title = "Loading " + str(num_caps) + " " + reftype + " item"
            title += "s" if len(capsule_paths) > 1 else ""
            first_chars = "A " if inventoryid == 0 else "An "
            first_chars += reftype + " Properties "
            QtGui.QMessageBox.information(
                self, title, first_chars + "dialog will appear for each " +
                "Data Capsule. Review each new item's properties, then " +
                "click 'OK' to save it or 'Cancel' to skip it and move onto " +
                "the next item.")
        for cap_path in capsule_paths:
            try:
                with open(cap_path, "r") as cap_file:
                    cap_data = json.load(cap_file)
                    if cap_data[0] != reftype:
                        QtGui.QMessageBox.warning(
                            self, "Invalid Data Capsule", "This Data " +
                            "Capsule does not contain " + reftype +
                            " reference item properties!")
                        continue
                    refobject = body.Reference(reftype, data_capsule=cap_data)
                    dialog = getattr(ears, "Input" + reftype)(
                        self, self.user, refobject)
                    if dialog.exec_():
                        # Save the reference and check for max ref IDs reached.
                        refobject = getattr(dialog, reftype.lower())
                        rid = self.user.add_reference(refobject)
                        if not rid:
                            QtGui.QMessageBox.warning(
                                self, "Maximum References Reached", "You " +
                                "cannot save anymore " + reftype + " items!")
                            return
                        # Reload inventory.
                        self.load_inventory(
                            inventoryid, keep_order=True,
                            selected_row=selectedrow)
                        QtGui.QMessageBox.information(
                            self, "Reference Item Created", "This " + reftype +
                            " item has been created:\n\n" + rid + ": " +
                            refobject.info("description"))
            except (IndexError, KeyError, TypeError, ValueError):
                QtGui.QMessageBox.warning(
                    self, "Invalid Data Capsule",
                    "You fed it a bad Data Capsule file:\n\n" + cap_path)
            except IOError:
                QtGui.QMessageBox.warning(
                    self, "Invalid File", "A Data Capsule cannot be loaded " +
                    "from the file path:\n\n" + cap_path)

    def view_item(self):
        """
        View the selected inventory item's properties.

        Initializes a build element object for the selected inventory item.
        Sets the state of the build element if it is a template or record.
        Calls ViewItemDialog with the build element, inventory ID, and item ID
        to let the user see it's properties.
        """
        inventoryid = self.inventory_field.currentIndex()
        selectedrow = self.inventoryviewer.currentRow()
        itemid = str(self.inventoryviewer.item(selectedrow, 0).text())
        cid = INVENTORY_CIDS[inventoryid]
        if inventoryid in range(2):
            item_element = body.build_element(cid, itemid)
        else:
            if inventoryid in range(2, 8):
                state_method = "template_state"
            else:
                state_method = "record_state"
            item_state = getattr(self.user, state_method)(cid, itemid)
            item_element = body.build_element(cid)
            item_element.set_state(item_state)
            item_element.widgetize()
        dialog = organs.ViewItemDialog(self, item_element, inventoryid, itemid)
        if dialog.exec_():
            pass

    def edit_item(self, inventoryid, itemid, selectedrow):
        """
        Edit the selected inventory item.

        Calls the applicable input dialog with a new Reference object or the
        corresponding template or record object to let the user edit its
        properties. If the dialog is accepted, calls the user's applicable
        add method with the Reference or build element object. Reloads the
        current inventory.

        :param inventoryid: The index of the currently selected inventory
        :type inventoryid: int
        :param itemid: The item ID of the selected inventory item
        :type itemid: str
        :param selectedrow: The selected row in the Inventory Viewer table
        :type selectedrow: int
        """
        cid = INVENTORY_CIDS[inventoryid]
        is_fave = True if itemid in self.user.settings()["favorites"][
            INVENTORY_FAVE_IDS[inventoryid]] else False
        if inventoryid in [0, 1]:
            reftype = "Food" if inventoryid == 0 else "Exercise"
            ref = body.Reference(reftype, item_id=itemid, user=self.user)
            dialog = getattr(ears, "Input" + reftype)(self, self.user, ref)
            if dialog.exec_():
                edited_ref = getattr(dialog, reftype.lower())
                self.user.add_reference(edited_ref)
                self.load_inventory(
                    inventoryid, keep_order=True, selected_row=selectedrow)
        elif inventoryid in range(2, 8):
            # Use the template object. It will be replaced in the template
            # objects dict if it is changed.
            element = self.user.template_objects(cid, itemid)
            dialog = ears.InputBuildElement(
                self, element, is_item=True, is_favorite=is_fave)
            if dialog.exec_():
                self.user.add_template(
                    element.template_state(), is_favorite=dialog.isfavorite,
                    item_id=itemid)
                self.load_inventory(
                    inventoryid, keep_order=True, selected_row=selectedrow)
        else:
            # Use the record object. It will be replaced in the record objects
            # dict if it is changed.
            element = self.user.record_objects(cid, itemid)
            old_date = self.user.record_state(cid, itemid)[3]
            dialog = ears.InputBuildElement(
                self, element, is_item=True, is_favorite=is_fave)
            if dialog.exec_():
                new_state = element.state()
                new_date = new_state[3]
                # Check if date is changed and a record exists with new date.
                if old_date != new_date and self.user.record_exists(new_state):
                    choice = QtGui.QMessageBox.warning(
                        self, "Confirm Overwrite Record",
                        "It looks like you changed this record's date, but " +
                        "another record already exists for this date. Do " +
                        "you want to replace the existing record with this " +
                        "one?", QtGui.QMessageBox.Yes | QtGui.QMessageBox.No)
                    if choice == QtGui.QMessageBox.No:
                        return
                # Check if a Program's start date is changed and it occurs
                # after it's earliest constituent Workout.
                if old_date != new_date and cid == "P":
                    earliest_period = sorted(
                        element.workout_periods().keys())[0]
                    earliest_date = str(datetime.datetime.strptime(
                        earliest_period, "%Y-%m-%d %H:%M").date())
                    if earliest_date < new_date:
                        choice = QtGui.QMessageBox.warning(
                            self, "Invalid Start Date",
                            "It looks like you changed the Program start " +
                            "date, but this new date occurs after the date " +
                            "of the earliest Workout. Your Program must " +
                            "start on or before Workout date:\n\n" +
                            str(earliest_date))
                        return
                # Pass old state to add_record if date/start date is changed.
                if old_date != new_date:
                    oldstate = self.user.record_state(cid, itemid)
                else:
                    oldstate = None
                self.user.add_record(
                    new_state, is_favorite=dialog.isfavorite,
                    old_state=oldstate)
                self.load_inventory(
                    inventoryid, keep_order=True, selected_row=selectedrow)

    def save_data_capsule(self, inventoryid, itemid):
        """
        Save the selected reference item as a data capsule.

        Instantiates a Reference object for the selected reference item and
        calls EasyFileDialog to let the user input the data capsule's file name
        and select the save directory. If the dialog is accepted, the data
        capsule is saved as a JSON formatted file.

        :param inventoryid: The index of the currently selected inventory
        :type inventoryid: int
        :param itemid: The item ID of the selected reference item
        :type itemid: str
        """
        try:
            reftype = "Food" if inventoryid == 0 else "Exercise"
            reference = body.Reference(reftype, itemid, self.user)
            datacapsule = reference.datacapsule()
            cap_name = reftype + "DataCapsule-" + reference.info("description")
            dialog = organs.EasyFileDialog(self, "savecapsule", cap_name)
            if dialog.exec_():
                dest_path = os.path.normpath(str(dialog.selectedFiles()[0]))
                if not dest_path[-5:] == ".json":
                    dest_path += ".json"
                if os.path.isfile(dest_path):
                    choice = QtGui.QMessageBox.warning(
                        self, "Confirm Overwrite File", "This file already " +
                        "exists! Do you want to replace it?",
                        QtGui.QMessageBox.Yes | QtGui.QMessageBox.No)
                    if choice == QtGui.QMessageBox.No:
                        return
                with open(dest_path, "w") as open_file:
                    ujson.dump(reference.datacapsule(), open_file)
                QtGui.QMessageBox.information(
                    self, "Data Capsule Saved",
                    "Your data capsule has been saved!")
        except IOError:
            QtGui.QMessageBox.warning(
                self, "File Error", "An error occurred while trying to save " +
                "this data capsule! Make sure you are not trying to save " +
                "over an open file.")

    def delete_item(self, inventoryid, itemid, selectedrow):
        """
        Delete the selected inventory item.

        Prompts the user to confirm item deletion and, if approved, calls the
        applicable GUI or user method to remove the item from the inventory. If
        the item is a reference, its use in the Build Parent, template items,
        and record items is assessed before deletion. If it is used in any way,
        the user is informed of its uses and the item is not deleted.

        :param inventoryid: The index of the currently selected inventory
        :type inventoryid: int
        :param itemid: The item ID of the selected inventory item
        :type itemid: str
        :param selectedrow: The selected row in the Inventory Viewer table
        :type selectedrow: int
        """
        cid = INVENTORY_CIDS[inventoryid]
        if inventoryid in [0, 1]:
            # Start a processor thread to look up item ID uses.
            self.processor = organs.DataProcessor(
                self, organs.process_item_counts)
            self.processor.data_processed.connect(self.delete_reference)
            reftype = "Food" if inventoryid == 0 else "Exercise"
            self.processor.set_kwargs(
                user=self.user, ref_type=reftype, item_id=itemid,
                item_row=selectedrow,
                build_parent=self.buildviewer.buildparent())
            self.proc_msg = organs.MessageDialog(
                self, "Searching Inventories", "LOOKING UP USES...")
            self.proc_msg.show()
            QtGui.QApplication.processEvents()
            self.processor.start()
        elif inventoryid in range(2, 8):
            if self.user.settings()["askdelete"]:
                choice = QtGui.QMessageBox.warning(
                    self, "Confirm Delete Template",
                    "Are you sure you want to delete this template item:\n\n" +
                    itemid + ": " + self.user.template_state(cid, itemid)[2],
                    QtGui.QMessageBox.Yes | QtGui.QMessageBox.No)
                if choice == QtGui.QMessageBox.No:
                    return
            self.user.remove_template(cid, itemid)
            self.load_inventory(
                inventoryid, keep_order=True, selected_row=selectedrow)
        elif inventoryid in [8, 9]:
            record_state = self.user.record_state(cid, itemid)
            if self.user.settings()["askdelete"]:
                choice = QtGui.QMessageBox.warning(
                    self, "Confirm Delete Record",
                    "Are you sure you want to delete this record item:\n\n" +
                    itemid + ": " + record_state[2],
                    QtGui.QMessageBox.Yes | QtGui.QMessageBox.No)
                if choice == QtGui.QMessageBox.No:
                    return
            self.user.remove_record(record_state)
            self.load_inventory(
                inventoryid, keep_order=True, selected_row=selectedrow)
            self.toggle_plot()

    def delete_reference(self, reference_counts):
        """
        Handles deletion of reference inventory items.

        This method is called by the DataProcessor thread started by the
        delete_item method to determine all uses of a reference item before it
        is deleted. Parses the data set emitted by the finished thread and
        deletes the reference item if it is not used in the Build Parent or by
        a template or record item. If uses are found, the item cannot be
        deleted. The user is informed of the outcome and, if deleted, the
        inventory is reloaded. The reference_counts arg emitted by the
        DataProcessor thread is in the format:

        (reference type, item ID, item row, in-buildparent bool, record count,
         template counts dict)

        :param reference_counts: Item count details in the format: (reference
            type, item ID, item row, in-buildparent bool, record count,
            template counts dict)
        :type reference_counts: tuple
        """
        reftype, itemid, itemrow, inbp, reccnt, temcnts = reference_counts
        self.proc_msg.finish()
        det_meth = "food_details" if reftype == "Food" else "exercise_details"
        desc = getattr(self.user, det_meth)(itemid)[0]
        if sum([reccnt] + temcnts.values()) == 0 and not inbp:
            if self.user.settings()["askdelete"]:
                choice = QtGui.QMessageBox.warning(
                    self, "Confirm Delete " + reftype,
                    "Are you sure you want to delete this " + reftype +
                    " item:\n\n" + itemid + ": " + desc,
                    QtGui.QMessageBox.Yes | QtGui.QMessageBox.No)
                if choice == QtGui.QMessageBox.No:
                    return
            ref = body.Reference(reftype, item_id=itemid, user=self.user)
            self.user.remove_reference(ref)
            inventoryid = self.inventory_field.currentIndex()
            self.load_inventory(
                inventoryid, keep_order=True, selected_row=itemrow)
            return
        # Case where item is in Build Parent.
        if inbp:
            QtGui.QMessageBox.warning(
                self, "Cannot Delete Item", "This " + name + " is currently " +
                "being used in your build! It cannot be deleted.")
            return
        # Case where item is in any records and/or templates.
        tem_cids = ["R", "M", "D"] if reftype == "Food" else ["W", "C", "P"]
        last_pos_tem_idx = 0
        for idx in range(3):
            if temcnts[tem_cids[idx]] > 0:
                last_pos_tem_idx = idx
        result = "".join([
            "This " + reftype + " cannot be deleted! It is currently being ",
            "used in:\n\n"])
        if reccnt > 0:
            result += str(reccnt) + " record"
            result += "s" if reccnt > 1 else ""
        first = True
        for idx in range(3):
            tem_cnt = temcnts[tem_cids[idx]]
            if tem_cnt > 0:
                tem_name = dna.BUILD_ELEMENTS[tem_cids[idx]][0]
                if first:
                    if reccnt > 0:
                        result += " and " if idx == last_pos_tem_idx else ", "
                    first = False
                else:
                    result += " and " if idx == last_pos_tem_idx else ", "
                result += str(tem_cnt) + " " + tem_name + " template"
                result += "s" if tem_cnt > 1 else ""
            if idx == last_pos_tem_idx:
                break
        QtGui.QMessageBox.warning(self, "Cannot Delete Item", result)

    def toggle_inventory(self):
        """
        Toggle Inventory Menu.

        Enables or disables Inventory Menu buttons based on the inventory_field
        attribute, the selected inventory item, the Build Parent, and the
        view_favorites attribute.
        """
        inventoryid = self.inventory_field.currentIndex()
        selected_row = self.inventoryviewer.currentRow()
        enable = []
        disable = []
        # Update Selected Item Viewer.
        if selected_row != -1:
            itemid = str(self.inventoryviewer.item(selected_row, 0).text())
            cid = INVENTORY_CIDS[inventoryid]
            selected_text = itemid + " - " + cid + ": "
            if inventoryid == 0:
                selected_text += self.user.food_details(itemid)[0]
            elif inventoryid == 1:
                selected_text += self.user.exercise_details(itemid)[0]
            elif inventoryid in range(2, 8):
                selected_text += self.user.template_state(cid, itemid)[2]
            else:
                selected_text += self.user.record_state(cid, itemid)[2]
        else:
            selected_text = ""
        self.selectedviewer.setText(str(selected_text))
        # Enable Send Item to Build if template or record is selected and there
        # is no Build Parent set.
        action = enable if (
            selected_row != -1 and self.buildviewer.buildparent() is None and
            inventoryid in range(2, 10)) else disable
        action.append(self.inventory_bns[0])
        # Enable Create New Reference and Load Data Capsule if ref inv loaded.
        action = enable if inventoryid in range(2) else disable
        action.extend([self.inventory_bns[idx] for idx in [1, 2]])
        # Enable View Item, Edit Item, and Delete Item if item selected.
        action = enable if selected_row != -1 else disable
        action.extend([self.inventory_bns[idx] for idx in [3, 4, 6]])
        # Enable Save Data Capsule if ref item selected.
        action = enable if (
            selected_row != -1 and inventoryid in range(2)) else disable
        action.append(self.inventory_bns[5])
        # Enable View All Items if viewing favorite items.
        action = enable if self.view_favorites else disable
        action.append(self.inventory_bns[7])
        # Enable View Favorite Items if viewing all items.
        action = enable if not self.view_favorites else disable
        action.append(self.inventory_bns[8])
        # Enable and disable buttons.
        for button in enable:
            button.setEnabled(True)
        for button in disable:
            button.setEnabled(False)

    # ------------------------ GUI METHODS ------------------------ #
    def center(self):
        """Center the window depending on the screen resolution."""
        screen = QtGui.QDesktopWidget().availableGeometry().center()
        if screen.y() > 500:
            # 1920 x 1080 and greater resolution.
            x = screen.x() - (self.width() / 2)
            y = screen.y() - (self.height() / 2)
            self.move(x, y)
        else:
            x = screen.x() - (self.width() / 2)
            self.move(x, 0)

    def closeEvent(self, event):
        """
        Close the application window.

        Overrides the inherited closeEvent method to prompt the user to confirm
        exiting the application. If the user's 'askexit' setting is False, this
        confirmation is skipped and the application is closed.

        :param event: Close event
        :type event: QCloseEvent
        """
        if self.user is not None and self.user.settings()["askexit"]:
            choice = QtGui.QMessageBox.warning(
                self, "Exit OnTrack",
                "Do you want to exit OnTrack and lose any unsaved data?",
                QtGui.QMessageBox.Yes | QtGui.QMessageBox.No)
            if choice == QtGui.QMessageBox.No:
                event.ignore()
            else:
                event.accept()
        else:
            event.accept()


def main():
    """Run the application."""
    app = QtGui.QApplication(sys.argv)
    gui = MainWindow()
    gui.show()
    sys.exit(app.exec_())


if __name__ == "__main__":
    # Create a lock file to prevent more than one application window from
    # running at the same time. If the reference source path does not exist,
    # exit the app.
    if not os.path.isdir(os.path.join(sys.path[0], dna.REF_DIR)):
        sys.exit(1)
    lockfile = os.path.join(sys.path[0], dna.REF_DIR, "locked.txt")
    try:
        if os.path.isfile(lockfile):
            os.unlink(lockfile)
    except WindowsError:
        sys.exit(0)
    # Open lock file and call main so that subsequent attempts to run an app
    # instance fail when they try to unlink the lock file.
    with open(lockfile, "wb") as temp_lock:
        main()
